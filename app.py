import sys
if sys.stdout.encoding and sys.stdout.encoding.lower() not in ('utf-8', 'utf8'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
if sys.stderr.encoding and sys.stderr.encoding.lower() not in ('utf-8', 'utf8'):
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file
import base64
import hashlib
from werkzeug.security import generate_password_hash, check_password_hash
import uuid
import threading
import smtplib
import json
import os
import calendar
from datetime import datetime, timedelta
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from io import BytesIO
from functools import wraps

try:
    import email_config
except ImportError:
    class email_config:
        SMTP_SERVER   = os.environ.get('SMTP_SERVER',   'smtp.gmail.com')
        SMTP_PORT     = int(os.environ.get('SMTP_PORT', '587'))
        SENDER_EMAIL  = os.environ.get('SMTP_EMAIL',    '')
        SENDER_PASSWORD = os.environ.get('SMTP_PASSWORD', '')
        ENABLED       = bool(SENDER_EMAIL and SENDER_PASSWORD)

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import qrcode
    from PIL import Image, ImageDraw, ImageFont
    HAS_QRCODE = True
except ImportError:
    HAS_QRCODE = False

# ── DB 백엔드 선택 ─────────────────────────────────────────────────────────────
DATABASE_URL = os.environ.get('DATABASE_URL', '')
if DATABASE_URL.startswith('postgres://'):
    DATABASE_URL = DATABASE_URL.replace('postgres://', 'postgresql://', 1)
USE_PG = bool(DATABASE_URL)

import sqlite3  # 항상 import (PostgreSQL 연결 실패 시 폴백용)

if USE_PG:
    try:
        import psycopg2
        import psycopg2.extras
        print(f"[DB] PostgreSQL 사용 (데이터 영구 보존)", flush=True)
    except ImportError:
        print("[DB] ⚠ psycopg2 없음 → SQLite 폴백", flush=True)
        USE_PG = False
else:
    print("[DB] ⚠ SQLite 사용 중 - DATABASE_URL 미설정. Render 재배포 시 데이터 초기화됨!", flush=True)

# SQL 방언 상수
_PK           = 'SERIAL PRIMARY KEY'          if USE_PG else 'INTEGER PRIMARY KEY AUTOINCREMENT'
_NOW_DEFAULT  = "to_char(NOW() AT TIME ZONE 'Asia/Seoul','YYYY-MM-DD HH24:MI:SS')" \
                if USE_PG else "datetime('now','localtime')"
TODAY         = 'CURRENT_DATE'                if USE_PG else "date('now','localtime')"
NOW_FN        = 'NOW()'                       if USE_PG else "datetime('now','localtime')"

def now_kst():
    """서버(UTC)와 무관하게 항상 한국 현재 시각(KST = UTC+9) 반환"""
    return datetime.utcnow() + timedelta(hours=9)

def date_col(col):
    return f"LEFT({col}, 10)" if USE_PG else f"date({col})"


# ── DB 연결 래퍼 ──────────────────────────────────────────────────────────────
class _PGRow(dict):
    """dict 를 sqlite3.Row 처럼 속성 접근도 지원하도록 확장"""
    def __getattr__(self, item):
        try:
            return self[item]
        except KeyError:
            raise AttributeError(item)


class _PGCursorWrapper:
    def __init__(self, cur):
        self._c = cur

    def _wrap(self, row):
        return _PGRow(row) if row is not None else None

    def fetchall(self):
        return [_PGRow(r) for r in self._c.fetchall()]

    def fetchone(self):
        row = self._c.fetchone()
        return _PGRow(row) if row else None

    def __iter__(self):
        return iter(self.fetchall())


class DBConn:
    def __init__(self):
        self._pg = USE_PG
        if USE_PG:
            try:
                self._conn = psycopg2.connect(DATABASE_URL)
                print('[DB] PostgreSQL 연결 성공', flush=True)
            except Exception as e:
                print(f'[DB] ⚠ PostgreSQL 연결 실패({e}) → SQLite 폴백', flush=True)
                self._pg = False
                self._conn = sqlite3.connect('facility.db')
                self._conn.row_factory = sqlite3.Row
        else:
            self._conn = sqlite3.connect('facility.db')
            self._conn.row_factory = sqlite3.Row

    def execute(self, sql, params=()):
        if self._pg:
            cur = self._conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            cur.execute(sql.replace('?', '%s'), params or None)
            return _PGCursorWrapper(cur)
        return self._conn.execute(sql, params)

    def insert(self, sql, params=()):
        """INSERT 후 새 row ID 반환"""
        if self._pg:
            cur = self._conn.cursor()
            pg_sql = (sql + ' RETURNING id').replace('?', '%s')
            cur.execute(pg_sql, params or None)
            return cur.fetchone()[0]
        cur = self._conn.execute(sql, params)
        return cur.lastrowid

    def commit(self):   self._conn.commit()
    def rollback(self): self._conn.rollback()
    def close(self):    self._conn.close()

    # 연결 타입에 맞는 SQL 방언 헬퍼
    def date_col(self, col):
        return f"LEFT({col}, 10)" if self._pg else f"date({col})"

    @property
    def today(self):
        return "to_char(NOW() AT TIME ZONE 'Asia/Seoul','YYYY-MM-DD')" if self._pg else "date('now','localtime')"

    @property
    def now_fn(self):
        return 'NOW()' if self._pg else "datetime('now','localtime')"


def get_db():
    return DBConn()


# ── Flask 앱 ──────────────────────────────────────────────────────────────────
app = Flask(__name__)
# ── 보안: SECRET_KEY 환경변수 강제화 (없으면 랜덤 생성 + 경고)
_secret_key = os.environ.get('SECRET_KEY')
if not _secret_key:
    import secrets as _secrets
    _secret_key = _secrets.token_hex(32)
    print('[보안경고] SECRET_KEY 환경변수 미설정 - 서버 재시작 시 세션이 초기화됩니다. Render 환경변수에 SECRET_KEY를 반드시 설정하세요.', flush=True)
app.secret_key = _secret_key
# ── 보안: 세션 30분 타임아웃
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(minutes=30)

TEAMS = ['품질팀', 'EMS제조팀', '생산기술팀', '개발팀', '환경안전팀', '사출팀', '코팅팀', '관리자']


def _dept_filter(conn, alias='e'):
    """URL 파라미터·세션 팀을 읽어 (WHERE절 조각, params리스트, 현재dept값) 반환.
    team='관리자' 또는 dept='전체' → 필터 없음.
    관리자도 명시적으로 팀 선택 가능.
    """
    dept = request.args.get('dept', '').strip()
    if not dept:
        # 기본값: 팀이 '관리자'가 아니고 관리자 계정도 아니면 자기 팀
        if not session.get('is_admin') and session.get('team') and session.get('team') != '관리자':
            dept = session.get('team', '')
        else:
            dept = '전체'
    if dept == '전체':
        return '', [], dept
    ph = '%s' if (USE_PG and conn._pg) else '?'
    return f' AND {alias}.department = {ph}', [dept], dept


# ── 엑셀 파싱 ─────────────────────────────────────────────────────────────────
def parse_excel(file):
    wb = openpyxl.load_workbook(file, data_only=True)
    ws = wb.active
    rows_data = []
    max_cols  = 0
    for row in ws.iter_rows(values_only=True):
        cells = []
        for c in row:
            if c is None:
                cells.append('')
            elif isinstance(c, float) and c == int(c):
                cells.append(str(int(c)))
            else:
                cells.append(str(c).strip())
        while cells and not cells[-1]:
            cells.pop()
        if not any(cells):
            continue
        max_cols = max(max_cols, len(cells))
        try:
            int(float(cells[0]))
            is_item = True
        except (ValueError, TypeError, IndexError):
            is_item = False
        rows_data.append({'cells': cells, 'is_item': is_item})
    return rows_data, max_cols


# ── 이메일 발송 ───────────────────────────────────────────────────────────────
def _build_email_html(approver_name, inspector_name, eq_name, location,
                      result, notes, inspect_url):
    result_color = {
        '정상': '#16a34a', '이상': '#f97316',
        '수리필요': '#dc2626', '휴동': '#6b7280',
    }.get(result, '#f97316')
    result_icon = {'정상': '✅', '이상': '⚠️', '수리필요': '🔴', '휴동': '⏸'}.get(result, '📋')
    notes_row = f'''
        <tr>
          <td style="padding:8px 0;color:#6b7280;width:100px;">특이사항</td>
          <td style="padding:8px 0;color:#111827;">{notes}</td>
        </tr>''' if notes else ''

    return f'''<!DOCTYPE html>
<html lang="ko"><head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;background:#f8f9fa;font-family:'Malgun Gothic',sans-serif;">
  <div style="max-width:520px;margin:32px auto;background:#fff;
              border-radius:16px;overflow:hidden;border:1px solid #e5e7eb;">
    <div style="background:linear-gradient(135deg,#f97316,#ea580c);
                padding:28px 32px;text-align:center;">
      <div style="color:#fff;font-size:1.2rem;font-weight:700;letter-spacing:1px;">
        INTOPS 설비점검 시스템
      </div>
      <div style="color:rgba(255,255,255,0.85);font-size:0.88rem;margin-top:4px;">
        점검 승인 요청 알림
      </div>
    </div>
    <div style="padding:32px;">
      <p style="color:#111827;font-size:1rem;margin:0 0 24px;">
        안녕하세요, <strong style="color:#f97316;">{approver_name}</strong> 님.<br>
        아래 설비에 대한 점검이 완료되어 승인 요청이 접수되었습니다.
      </p>
      <div style="background:#f8f9fa;border-radius:10px;padding:20px 24px;
                  border:1px solid #e5e7eb;margin-bottom:24px;">
        <table style="width:100%;border-collapse:collapse;">
          <tr><td style="padding:8px 0;color:#6b7280;width:100px;">설비명</td>
              <td style="padding:8px 0;color:#111827;font-weight:700;">{eq_name}</td></tr>
          <tr><td style="padding:8px 0;color:#6b7280;">설치 위치</td>
              <td style="padding:8px 0;color:#111827;">{location}</td></tr>
          <tr><td style="padding:8px 0;color:#6b7280;">점검자</td>
              <td style="padding:8px 0;color:#111827;">{inspector_name}</td></tr>
          <tr><td style="padding:8px 0;color:#6b7280;">점검 결과</td>
              <td style="padding:8px 0;">
                <span style="background:{result_color}18;color:{result_color};
                             padding:3px 12px;border-radius:8px;font-weight:700;">
                  {result_icon} {result}
                </span>
              </td></tr>
          {notes_row}
        </table>
      </div>
      <div style="text-align:center;margin-bottom:24px;">
        <a href="{inspect_url}"
           style="display:inline-block;background:linear-gradient(135deg,#f97316,#ea580c);
                  color:#fff;text-decoration:none;padding:14px 36px;
                  border-radius:10px;font-weight:700;">
          승인 확인하기 →
        </a>
      </div>
      <p style="color:#9ca3af;font-size:0.8rem;text-align:center;margin:0;">
        이 메일은 INTOPS 설비점검 시스템에서 자동 발송된 메일입니다.
      </p>
    </div>
  </div>
</body></html>'''


def _send_mail(to_email, subject, html_body):
    try:
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From']    = email_config.SENDER_EMAIL
        msg['To']      = to_email
        msg.attach(MIMEText(html_body, 'html', 'utf-8'))
        with smtplib.SMTP(email_config.SMTP_SERVER, email_config.SMTP_PORT) as s:
            s.starttls()
            s.login(email_config.SENDER_EMAIL, email_config.SENDER_PASSWORD)
            s.sendmail(email_config.SENDER_EMAIL, to_email, msg.as_string())
        print(f'[이메일] 발송 완료 → {to_email}')
        return True
    except Exception as e:
        print(f'[이메일] 발송 실패: {e}', flush=True)
        import sys; sys.stderr.write(f'[이메일] 발송 실패: {e}\n'); sys.stderr.flush()
        return str(e)  # 에러 메시지 반환


def _auto_fill_cycle(conn, eq_id, inspector_id, approver_id, result, inspected_date_str, cycle):
    """주1회/월1회 점검 시 해당 기간 나머지 평일에 자동 승인완료 기록 삽입."""
    if cycle not in ('주1회', '월1회'):
        return 0
    ph = '%s' if conn._pg else '?'
    day_col = "LEFT(inspected_at,10)" if conn._pg else "substr(inspected_at,1,10)"
    base = datetime.strptime(inspected_date_str, '%Y-%m-%d')

    if cycle == '주1회':
        mon = base - timedelta(days=base.weekday())
        target_dates = [mon + timedelta(days=i) for i in range(5)]  # 월~금
    else:  # 월1회
        days_in_m = calendar.monthrange(base.year, base.month)[1]
        target_dates = [
            datetime(base.year, base.month, d)
            for d in range(1, days_in_m + 1)
            if datetime(base.year, base.month, d).weekday() < 5
        ]

    auto_notes = f'자동입력({cycle})'
    inserted   = 0
    for d in target_dates:
        if d.date() == base.date():
            continue                    # 점검한 날은 이미 존재
        ds = d.strftime('%Y-%m-%d')
        if conn.execute(f'SELECT id FROM inspections WHERE equipment_id={ph} AND {day_col}={ph}',
                        (eq_id, ds)).fetchone():
            continue                    # 이미 기록 있는 날 건너뜀
        ts = ds + ' 00:00:00'
        app_id = approver_id or inspector_id
        if conn._pg:
            conn.execute('''INSERT INTO inspections
                (equipment_id,inspector_id,result,notes,status,inspected_at,approved_by,approved_at)
                VALUES (%s,%s,%s,%s,'승인완료',%s,%s,NOW())''',
                (eq_id, inspector_id, result, auto_notes, ts, app_id))
        else:
            conn.execute('''INSERT INTO inspections
                (equipment_id,inspector_id,result,notes,status,inspected_at,approved_by,approved_at)
                VALUES (?,?,?,?,'승인완료',?,?,datetime('now','localtime'))''',
                (eq_id, inspector_id, result, auto_notes, ts, app_id))
        inserted += 1
    return inserted


def send_approval_request(to_email, approver_name, inspector_name,
                          eq_name, location, result, notes, eq_id, host_url):
    if not email_config.ENABLED or not to_email:
        return
    if get_setting('email_enabled', '1') != '1':
        return
    inspect_url = host_url.rstrip('/') + url_for('inspect', eq_id=eq_id)
    subject     = f'[설비점검] 승인 요청 - {eq_name}'
    html_body   = _build_email_html(approver_name, inspector_name,
                                    eq_name, location, result, notes, inspect_url)
    t = threading.Thread(target=_send_mail, args=(to_email, subject, html_body), daemon=True)
    t.start()


# ── DB 초기화 ─────────────────────────────────────────────────────────────────
def init_db():
    conn = get_db()
    # 실제 연결 타입(PG vs SQLite) 기반으로 SQL 방언 결정
    _pk  = 'SERIAL PRIMARY KEY' if conn._pg else 'INTEGER PRIMARY KEY AUTOINCREMENT'
    _now = ("to_char(NOW() AT TIME ZONE 'Asia/Seoul','YYYY-MM-DD HH24:MI:SS')"
            if conn._pg else "datetime('now','localtime')")

    conn.execute(f'''
        CREATE TABLE IF NOT EXISTS users (
            id          {_pk},
            name        TEXT NOT NULL,
            employee_id TEXT UNIQUE NOT NULL,
            email       TEXT DEFAULT '',
            phone       TEXT NOT NULL,
            team        TEXT NOT NULL,
            password    TEXT NOT NULL,
            role        TEXT DEFAULT '점검자',
            is_admin    INTEGER DEFAULT 0,
            is_approved INTEGER DEFAULT 0,
            created_at  TEXT DEFAULT ({_now})
        )
    ''')

    conn.execute(f'''
        CREATE TABLE IF NOT EXISTS equipment (
            id          {_pk},
            name        TEXT NOT NULL,
            qr_code     TEXT UNIQUE NOT NULL,
            location    TEXT,
            department  TEXT,
            description TEXT,
            approver_id INTEGER,
            created_by  INTEGER,
            created_at  TEXT DEFAULT ({_now})
        )
    ''')

    conn.execute(f'''
        CREATE TABLE IF NOT EXISTS inspections (
            id            {_pk},
            equipment_id  INTEGER NOT NULL,
            inspector_id  INTEGER NOT NULL,
            result        TEXT NOT NULL,
            notes         TEXT,
            status        TEXT DEFAULT '점검완료',
            approved_by   INTEGER,
            approved_at   TEXT,
            inspected_at  TEXT DEFAULT ({_now})
        )
    ''')

    conn.execute(f'''
        CREATE TABLE IF NOT EXISTS inspection_templates (
            id           {_pk},
            equipment_id INTEGER UNIQUE NOT NULL,
            filename     TEXT,
            max_cols     INTEGER DEFAULT 0,
            rows         TEXT,
            created_at   TEXT DEFAULT ({_now})
        )
    ''')

    conn.execute(f'''
        CREATE TABLE IF NOT EXISTS inspection_details (
            id            {_pk},
            inspection_id INTEGER NOT NULL,
            row_index     INTEGER NOT NULL,
            result        TEXT NOT NULL DEFAULT '정상',
            detail_notes  TEXT DEFAULT ''
        )
    ''')

    conn.execute(f'''
        CREATE TABLE IF NOT EXISTS inspection_items (
            id           {_pk},
            equipment_id INTEGER NOT NULL,
            item_order   INTEGER DEFAULT 0,
            category     TEXT DEFAULT '',
            item_name    TEXT NOT NULL,
            criteria     TEXT DEFAULT '',
            unit         TEXT DEFAULT '',
            item_type    TEXT DEFAULT '일반',
            min_val      TEXT DEFAULT '',
            center_val   TEXT DEFAULT '',
            max_val      TEXT DEFAULT '',
            created_at   TEXT DEFAULT ({_now})
        )
    ''')

    conn.execute(f'''
        CREATE TABLE IF NOT EXISTS monthly_notes (
            id           {_pk},
            equipment_id INTEGER NOT NULL,
            year         INTEGER NOT NULL,
            month        INTEGER NOT NULL,
            notes        TEXT DEFAULT '',
            updated_at   TEXT DEFAULT ({_now}),
            UNIQUE(equipment_id, year, month)
        )
    ''')

    conn.execute(f'''
        CREATE TABLE IF NOT EXISTS password_reset_requests (
            id           {_pk},
            user_id      INTEGER NOT NULL,
            status       TEXT DEFAULT '대기중',
            reset_code   TEXT DEFAULT '',
            reset_expires TEXT DEFAULT '',
            created_at   TEXT DEFAULT ({_now})
        )
    ''')

    conn.execute('''
        CREATE TABLE IF NOT EXISTS system_settings (
            key   TEXT PRIMARY KEY,
            value TEXT DEFAULT ''
        )
    ''')

    conn.execute(f'''
        CREATE TABLE IF NOT EXISTS equipment_anomalies (
            id            {_pk},
            equipment_id  INTEGER NOT NULL,
            inspection_id INTEGER,
            reporter_id   INTEGER NOT NULL,
            occurred_at   TEXT DEFAULT ({_now}),
            description   TEXT NOT NULL,
            action_taken  TEXT DEFAULT '',
            action_person TEXT DEFAULT '',
            priority      TEXT DEFAULT '보통',
            is_resolved   INTEGER DEFAULT 0,
            resolved_at   TEXT,
            resolved_by   INTEGER,
            created_at    TEXT DEFAULT ({_now})
        )
    ''')

    conn.execute(f'''
        CREATE TABLE IF NOT EXISTS anomaly_photos (
            id          {_pk},
            anomaly_id  INTEGER NOT NULL,
            photo_data  TEXT NOT NULL,
            filename    TEXT DEFAULT '',
            created_at  TEXT DEFAULT ({_now})
        )
    ''')

    # 마이그레이션
    if conn._pg:
        migrations = [
            "ALTER TABLE users ADD COLUMN IF NOT EXISTS email TEXT DEFAULT ''",
            "ALTER TABLE users ADD COLUMN IF NOT EXISTS role TEXT DEFAULT '점검자'",
            "ALTER TABLE equipment ADD COLUMN IF NOT EXISTS approver_id INTEGER",
            "ALTER TABLE inspections ADD COLUMN IF NOT EXISTS status TEXT DEFAULT '점검완료'",
            "ALTER TABLE inspections ADD COLUMN IF NOT EXISTS approved_by INTEGER",
            "ALTER TABLE inspections ADD COLUMN IF NOT EXISTS approved_at TEXT",
            "ALTER TABLE equipment ADD COLUMN IF NOT EXISTS inspection_cycle TEXT DEFAULT '매일'",
            "ALTER TABLE inspection_details ADD COLUMN IF NOT EXISTS item_id INTEGER",
            "ALTER TABLE equipment ADD COLUMN IF NOT EXISTS mgmt_no TEXT DEFAULT ''",
            "ALTER TABLE equipment ADD COLUMN IF NOT EXISTS manager_primary TEXT DEFAULT ''",
            "ALTER TABLE equipment ADD COLUMN IF NOT EXISTS manager_secondary TEXT DEFAULT ''",
            "ALTER TABLE password_reset_requests ADD COLUMN IF NOT EXISTS reset_code TEXT DEFAULT ''",
            "ALTER TABLE password_reset_requests ADD COLUMN IF NOT EXISTS reset_expires TEXT DEFAULT ''",
            "ALTER TABLE inspection_items ADD COLUMN IF NOT EXISTS item_type TEXT DEFAULT '일반'",
            "ALTER TABLE inspection_items ADD COLUMN IF NOT EXISTS min_val TEXT DEFAULT ''",
            "ALTER TABLE inspection_items ADD COLUMN IF NOT EXISTS center_val TEXT DEFAULT ''",
            "ALTER TABLE inspection_items ADD COLUMN IF NOT EXISTS max_val TEXT DEFAULT ''",
            # 기존 unit 텍스트값 정규화 (공백 제거)
            "UPDATE inspection_items SET unit='주1회' WHERE item_type='일반' AND REPLACE(unit,' ','')='주1회' AND unit<>'주1회'",
            "UPDATE inspection_items SET unit='월1회' WHERE item_type='일반' AND REPLACE(unit,' ','')='월1회' AND unit<>'월1회'",
            "UPDATE inspection_items SET unit='일1회' WHERE item_type='일반' AND unit IN ('매일','일 1회','일1회 ','daily','')",
            "ALTER TABLE equipment_anomalies ADD COLUMN IF NOT EXISTS action_person TEXT DEFAULT ''",
        ]
        for sql in migrations:
            try:
                conn.execute(sql)
            except Exception:
                conn._conn.rollback()
    else:
        migrations = [
            "ALTER TABLE users ADD COLUMN email TEXT DEFAULT ''",
            "ALTER TABLE users ADD COLUMN role TEXT DEFAULT '점검자'",
            "ALTER TABLE equipment ADD COLUMN approver_id INTEGER",
            "ALTER TABLE inspections ADD COLUMN status TEXT DEFAULT '점검완료'",
            "ALTER TABLE inspections ADD COLUMN approved_by INTEGER",
            "ALTER TABLE inspections ADD COLUMN approved_at TEXT",
            "ALTER TABLE equipment ADD COLUMN inspection_cycle TEXT DEFAULT '매일'",
            "ALTER TABLE inspection_details ADD COLUMN item_id INTEGER",
            "ALTER TABLE equipment ADD COLUMN mgmt_no TEXT DEFAULT ''",
            "ALTER TABLE equipment ADD COLUMN manager_primary TEXT DEFAULT ''",
            "ALTER TABLE equipment ADD COLUMN manager_secondary TEXT DEFAULT ''",
            "ALTER TABLE password_reset_requests ADD COLUMN reset_code TEXT DEFAULT ''",
            "ALTER TABLE password_reset_requests ADD COLUMN reset_expires TEXT DEFAULT ''",
            "ALTER TABLE inspection_items ADD COLUMN item_type TEXT DEFAULT '일반'",
            "ALTER TABLE inspection_items ADD COLUMN min_val TEXT DEFAULT ''",
            "ALTER TABLE inspection_items ADD COLUMN center_val TEXT DEFAULT ''",
            "ALTER TABLE inspection_items ADD COLUMN max_val TEXT DEFAULT ''",
            "UPDATE inspection_items SET unit='주1회' WHERE item_type='일반' AND REPLACE(unit,' ','')='주1회' AND unit<>'주1회'",
            "UPDATE inspection_items SET unit='월1회' WHERE item_type='일반' AND REPLACE(unit,' ','')='월1회' AND unit<>'월1회'",
            "UPDATE inspection_items SET unit='일1회' WHERE item_type='일반' AND unit IN ('매일','일 1회','일1회 ','daily','')",
            "ALTER TABLE equipment_anomalies ADD COLUMN action_person TEXT DEFAULT ''",
        ]
        for sql in migrations:
            try:
                conn.execute(sql)
            except Exception:
                pass

    # 관리자 초기 비밀번호: 환경변수 ADMIN_PASSWORD 우선, 없으면 admin123 (반드시 변경 권고)
    _admin_raw = os.environ.get('ADMIN_PASSWORD', 'admin123')
    if _admin_raw == 'admin123':
        print('[보안경고] ADMIN_PASSWORD 환경변수가 설정되지 않았습니다. 기본값(admin123)을 사용 중입니다. 즉시 변경하세요.', flush=True)
    admin_pw = generate_password_hash(_admin_raw)

    if conn._pg:
        conn.execute('''
            INSERT INTO users (name, employee_id, email, phone, team, password, role, is_admin, is_approved)
            VALUES (?, ?, ?, ?, ?, ?, ?, 1, 1)
            ON CONFLICT (employee_id) DO NOTHING
        ''', ('관리자', 'admin', 'admin@company.com', '010-0000-0000', '경영지원그룹', admin_pw, '승인자'))
    else:
        conn.execute('''
            INSERT OR IGNORE INTO users
                (name, employee_id, email, phone, team, password, role, is_admin, is_approved)
            VALUES (?, ?, ?, ?, ?, ?, ?, 1, 1)
        ''', ('관리자', 'admin', 'admin@company.com', '010-0000-0000', '경영지원그룹', admin_pw, '승인자'))

    conn.commit()
    conn.close()


# ── 시스템 설정 헬퍼 ──────────────────────────────────────────────────────────
def get_setting(key, default=''):
    """system_settings 테이블에서 값을 읽는다."""
    try:
        conn = get_db()
        ph   = '%s' if conn._pg else '?'
        row  = conn.execute(
            f'SELECT value FROM system_settings WHERE key = {ph}', (key,)
        ).fetchone()
        conn.close()
        return row['value'] if row else default
    except Exception:
        return default

def set_setting(key, value):
    """system_settings 테이블에 값을 저장(upsert)한다."""
    conn = get_db()
    if conn._pg:
        conn.execute(
            'INSERT INTO system_settings (key, value) VALUES (%s, %s) '
            'ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value',
            (key, str(value))
        )
    else:
        conn.execute(
            'INSERT OR REPLACE INTO system_settings (key, value) VALUES (?, ?)',
            (key, str(value))
        )
    conn.commit()
    conn.close()


def hash_pw(pw):
    # 신규 비밀번호: werkzeug pbkdf2 해시 사용 (SHA256 대비 보안 강화)
    return generate_password_hash(pw)

def check_pw(stored_hash, pw):
    # 구버전(SHA256) 호환: 기존 사용자 로그인 시 자동 마이그레이션
    if stored_hash.startswith('pbkdf2:') or stored_hash.startswith('scrypt:'):
        return check_password_hash(stored_hash, pw)
    # 구버전 SHA256 해시 확인
    import hashlib as _hl
    return _hl.sha256(pw.encode()).hexdigest() == stored_hash


# ── 비밀번호 재설정 인증코드 임시 저장소 ─────────────────────────────────────
# { email: {'code': '123456', 'user_id': 1, 'expires': datetime} }
_reset_store = {}
_reset_lock  = threading.Lock()

def _clean_expired_codes():
    now = datetime.now()
    with _reset_lock:
        expired = [e for e, v in _reset_store.items() if v['expires'] < now]
        for e in expired:
            del _reset_store[e]


def send_reset_code(to_email, user_name, code):
    subject = '[INTOPS] 비밀번호 재설정 인증번호'
    html = f'''<!DOCTYPE html><html><body style="margin:0;padding:0;background:#f9fafb;font-family:sans-serif;">
  <div style="max-width:480px;margin:40px auto;background:#fff;border-radius:16px;
              box-shadow:0 2px 16px rgba(0,0,0,0.08);overflow:hidden;">
    <div style="background:#f97316;padding:28px 32px;">
      <h2 style="color:#fff;margin:0;font-size:1.3rem;">🔐 비밀번호 재설정</h2>
    </div>
    <div style="padding:32px;">
      <p style="color:#374151;margin:0 0 12px;"><strong>{user_name}</strong> 님, 안녕하세요.</p>
      <p style="color:#6b7280;margin:0 0 24px;">아래 인증번호를 입력하여 비밀번호를 재설정하세요.<br>
         인증번호는 <strong>10분간</strong> 유효합니다.</p>
      <div style="text-align:center;background:#fff7ed;border:2px dashed #f97316;
                  border-radius:12px;padding:24px;margin:24px 0;">
        <span style="font-size:2.2rem;font-weight:900;letter-spacing:8px;color:#f97316;">{code}</span>
      </div>
      <p style="color:#9ca3af;font-size:0.8rem;margin:0;">
        본인이 요청하지 않은 경우 이 메일을 무시하세요.
      </p>
    </div>
  </div>
</body></html>'''
    threading.Thread(target=_send_mail, args=(to_email, subject, html), daemon=True).start()


# ── 평일 오전 11시 미점검 알림 ───────────────────────────────────────────────
def _send_inspection_reminders():
    """평일 오전 11시(KST), 당일 미점검 설비 담당자(정/부)에게 알림 이메일 발송"""
    if not email_config.ENABLED:
        print('[알림] SMTP 미설정 - 점검 알림 스킵', flush=True)
        return
    if get_setting('email_enabled', '1') != '1':
        print('[알림] 관리자 설정으로 이메일 발송 OFF - 점검 알림 스킵', flush=True)
        return

    now = datetime.now()
    today_str = now.strftime('%Y-%m-%d')
    print(f'[알림] 미점검 알림 실행 - {today_str}', flush=True)

    conn = get_db()
    try:
        # 당일 미점검 설비 목록
        if conn._pg:
            rows = conn.execute("""
                SELECT e.id, e.name, e.manager_primary, e.manager_secondary
                FROM equipment e
                WHERE NOT EXISTS (
                    SELECT 1 FROM inspections i
                    WHERE i.equipment_id = e.id
                    AND DATE(i.inspected_at) = %s
                )
            """, (today_str,)).fetchall()
        else:
            rows = conn.execute("""
                SELECT e.id, e.name, e.manager_primary, e.manager_secondary
                FROM equipment e
                WHERE NOT EXISTS (
                    SELECT 1 FROM inspections i
                    WHERE i.equipment_id = e.id
                    AND DATE(i.inspected_at) = ?
                )
            """, (today_str,)).fetchall()

        for eq in rows:
            recipients = set()
            for mgr_name in [eq['manager_primary'], eq['manager_secondary']]:
                if not mgr_name:
                    continue
                if conn._pg:
                    u = conn.execute(
                        "SELECT email FROM users WHERE TRIM(name)=%s AND is_approved=1",
                        (mgr_name.strip(),)
                    ).fetchone()
                else:
                    u = conn.execute(
                        "SELECT email FROM users WHERE TRIM(name)=? AND is_approved=1",
                        (mgr_name.strip(),)
                    ).fetchone()
                if u and u['email']:
                    recipients.add(u['email'])

            for email in recipients:
                subject = f'[INTOPS] 점검 알림 - {eq["name"]} 오늘 미점검'
                html = f'''<!DOCTYPE html><html><body style="margin:0;padding:0;background:#f9fafb;font-family:sans-serif;">
<div style="max-width:480px;margin:40px auto;background:#fff;border-radius:16px;
            box-shadow:0 2px 16px rgba(0,0,0,0.08);overflow:hidden;">
  <div style="background:#f97316;padding:24px 32px;">
    <h2 style="color:#fff;margin:0;font-size:1.2rem;">⚠️ 설비 점검 알림</h2>
  </div>
  <div style="padding:28px 32px;">
    <p style="color:#374151;margin:0 0 16px;">안녕하세요.</p>
    <p style="color:#374151;margin:0 0 16px;">
      오늘 <strong style="color:#f97316;">{now.strftime("%Y년 %m월 %d일")}</strong> 오전 11시까지
      아래 설비의 점검이 완료되지 않았습니다.
    </p>
    <div style="background:#fff7ed;border:2px solid #f97316;border-radius:10px;
                padding:16px 20px;margin:0 0 20px;">
      <div style="font-size:1.1rem;font-weight:800;color:#c2410c;">{eq["name"]}</div>
    </div>
    <p style="color:#6b7280;font-size:0.85rem;margin:0;">
      점검을 완료한 경우 이 알림을 무시하세요.
    </p>
  </div>
</div></body></html>'''
                _send_mail(email, subject, html)
                print(f'[알림] {eq["name"]} → {email} 발송', flush=True)
    finally:
        conn.close()


# gunicorn 포함 모든 실행 환경에서 DB 초기화 보장
with app.app_context():
    init_db()

    # 평일 오전 11시(KST) 미점검 알림 스케줄러
    try:
        from apscheduler.schedulers.background import BackgroundScheduler
        from apscheduler.triggers.cron import CronTrigger
        import pytz
        _kst = pytz.timezone('Asia/Seoul')
        _scheduler = BackgroundScheduler(timezone=_kst)
        _scheduler.add_job(
            _send_inspection_reminders,
            CronTrigger(day_of_week='mon-fri', hour=11, minute=0, timezone=_kst)
        )
        _scheduler.start()
        print('[스케줄러] 평일 오전 11시 미점검 알림 등록 완료', flush=True)
    except Exception as _e:
        print(f'[스케줄러] 시작 실패: {_e}', flush=True)

    # 시작 시 DB 연결 상태 명확히 출력
    _chk = get_db()
    if _chk._pg:
        print("✅ [DB] PostgreSQL 연결 성공 - 데이터 영구 보존됩니다.", flush=True)
    else:
        print("⚠️  [DB] SQLite 사용 중 - Render 재배포 시 데이터가 초기화됩니다!", flush=True)
        print(f"⚠️  [DB] DATABASE_URL 확인 필요: {DATABASE_URL[:40]}..." if DATABASE_URL else "⚠️  [DB] DATABASE_URL 미설정", flush=True)
    _chk.close()


# ── 데코레이터 ────────────────────────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get('user_id'):
            return redirect(url_for('login', next=request.url))
        return f(*args, **kwargs)
    return wrapper


def admin_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get('is_admin'):
            flash('관리자 권한이 필요합니다.', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return wrapper


# ── 스플래시 ──────────────────────────────────────────────────────────────────
@app.route('/')
def splash():
    return render_template('splash.html')


# ── 로그인 실패 횟수 추적 (IP 기반, 5회 실패 시 15분 잠금) ──────────────────
_login_fail_store = {}  # { ip: {'count': 0, 'locked_until': None} }
_login_fail_lock  = threading.Lock()
LOGIN_MAX_FAIL    = 5          # 최대 실패 허용 횟수
LOGIN_LOCK_MIN    = 15         # 잠금 시간(분)

def _check_login_lock(ip):
    """IP 잠금 여부 확인. 잠겨있으면 남은 시간(분) 반환, 아니면 None"""
    with _login_fail_lock:
        info = _login_fail_store.get(ip)
        if not info:
            return None
        if info.get('locked_until') and datetime.now() < info['locked_until']:
            remaining = int((info['locked_until'] - datetime.now()).total_seconds() / 60) + 1
            return remaining
        # 잠금 해제 또는 미잠금
        return None

def _record_login_fail(ip):
    """로그인 실패 기록. 5회 초과 시 15분 잠금"""
    with _login_fail_lock:
        info = _login_fail_store.setdefault(ip, {'count': 0, 'locked_until': None})
        info['count'] += 1
        if info['count'] >= LOGIN_MAX_FAIL:
            info['locked_until'] = datetime.now() + timedelta(minutes=LOGIN_LOCK_MIN)
            info['count'] = 0  # 잠금 후 카운터 초기화
            print(f'[보안] 로그인 {LOGIN_MAX_FAIL}회 실패 - IP {ip} {LOGIN_LOCK_MIN}분 잠금', flush=True)

def _reset_login_fail(ip):
    """로그인 성공 시 실패 기록 초기화"""
    with _login_fail_lock:
        _login_fail_store.pop(ip, None)

# ── 로그인 ────────────────────────────────────────────────────────────────────
@app.route('/login', methods=['GET', 'POST'])
def login():
    next_url = request.args.get('next', '')
    if request.method == 'POST':
        client_ip = request.remote_addr or 'unknown'
        emp_id    = request.form['employee_id'].strip()
        raw_pw    = request.form['password']
        next_url  = request.form.get('next', '')

        # IP 잠금 확인
        lock_remain = _check_login_lock(client_ip)
        if lock_remain:
            flash(f'로그인 시도가 너무 많습니다. {lock_remain}분 후 다시 시도하세요.', 'error')
            return render_template('login.html', next_url=next_url)

        conn = get_db()
        user = conn.execute(
            'SELECT * FROM users WHERE employee_id=?', (emp_id,)
        ).fetchone()

        if user and check_pw(user['password'], raw_pw):
            # 로그인 성공: SHA256 구버전 해시이면 pbkdf2로 자동 마이그레이션
            stored = user['password']
            if not (stored.startswith('pbkdf2:') or stored.startswith('scrypt:')):
                new_hash = hash_pw(raw_pw)
                conn.execute('UPDATE users SET password=? WHERE id=?', (new_hash, user['id']))
                conn.commit()
            conn.close()
            if not user['is_approved']:
                flash('관리자 승인 대기 중입니다.', 'warning')
                return render_template('login.html', next_url=next_url)
            _reset_login_fail(client_ip)
            session.permanent = True  # 세션 타임아웃 적용
            session['user_id']   = user['id']
            session['user_name'] = user['name']
            session['is_admin']  = bool(user['is_admin'])
            session['role']      = user['role'] or '점검자'
            session['team']      = user.get('team') or ''
            if next_url:
                return redirect(next_url)
            if user['is_admin']:
                return redirect(url_for('admin'))
            return redirect(url_for('dashboard'))
        else:
            conn.close()
            _record_login_fail(client_ip)
            flash('사번 또는 비밀번호가 올바르지 않습니다.', 'error')
    return render_template('login.html', next_url=next_url)


# ── 비밀번호 찾기 1단계: 사번 + 이름으로 본인 확인 ──────────────────────────
@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        emp_id = request.form.get('employee_id', '').strip()
        name   = request.form.get('name', '').strip()
        print(f'[비번찾기] 입력값: emp_id={emp_id!r}, name={name!r}', flush=True)
        conn   = get_db()
        user   = conn.execute(
            'SELECT id, name, email FROM users WHERE TRIM(employee_id)=? AND TRIM(name)=?',
            (emp_id.strip(), name.strip())
        ).fetchone()
        print(f'[비번찾기] 조회결과: {dict(user) if user else None}', flush=True)
        conn.close()

        if not user:
            flash('사번 또는 이름이 일치하는 계정이 없습니다.', 'error')
            return render_template('forgot_password.html')

        import random
        from datetime import timedelta
        code  = f'{random.randint(0, 999999):06d}'
        key   = emp_id  # store key로 사번 사용
        _clean_expired_codes()
        with _reset_lock:
            _reset_store[key] = {
                'code':    code,
                'user_id': user['id'],
                'expires': datetime.now() + timedelta(minutes=10),
            }

        mail_sent = False
        if not email_config.ENABLED:
            mail_sent = f'SMTP 환경변수 미설정 (EMAIL={email_config.SENDER_EMAIL!r})'
        elif get_setting('email_enabled', '1') != '1':
            mail_sent = '관리자 설정으로 이메일 발송 OFF'
        elif not user['email']:
            mail_sent = '계정에 이메일 주소가 없음'
        if email_config.ENABLED and get_setting('email_enabled', '1') == '1' and user['email']:
            # 동기 발송으로 성공 여부 즉시 확인
            subject = '[INTOPS] 비밀번호 재설정 인증번호'
            html = f'''<!DOCTYPE html><html><body style="margin:0;padding:0;background:#f9fafb;font-family:sans-serif;">
  <div style="max-width:480px;margin:40px auto;background:#fff;border-radius:16px;
              box-shadow:0 2px 16px rgba(0,0,0,0.08);overflow:hidden;">
    <div style="background:#f97316;padding:28px 32px;">
      <h2 style="color:#fff;margin:0;font-size:1.3rem;">비밀번호 재설정</h2>
    </div>
    <div style="padding:32px;">
      <p style="color:#374151;margin:0 0 12px;"><strong>{user["name"]}</strong> 님, 안녕하세요.</p>
      <p style="color:#6b7280;margin:0 0 24px;">아래 인증번호를 입력하여 비밀번호를 재설정하세요.<br>
         인증번호는 <strong>10분간</strong> 유효합니다.</p>
      <div style="text-align:center;background:#fff7ed;border:2px dashed #f97316;
                  border-radius:12px;padding:24px;margin:24px 0;">
        <span style="font-size:2.2rem;font-weight:900;letter-spacing:8px;color:#f97316;">{code}</span>
      </div>
      <p style="color:#9ca3af;font-size:0.8rem;margin:0;">
        본인이 요청하지 않은 경우 이 메일을 무시하세요.
      </p>
    </div>
  </div></body></html>'''
            mail_sent = _send_mail(user['email'], subject, html)

        if mail_sent is True:
            masked = user['email']
            if '@' in masked:
                local, domain = masked.split('@', 1)
                masked = local[:2] + '***@' + domain
            flash(f'{masked} 으로 인증번호를 발송했습니다. 10분 내 입력하세요.', 'success')
            return redirect(url_for('verify_reset_code', emp_id=key))
        else:
            # 이메일 발송 실패 or 미설정 → 관리자 요청 화면 (emp_id 전달)
            err_msg = mail_sent if isinstance(mail_sent, str) else '미설정'
            flash(f'이메일 발송 오류: {err_msg}', 'error')
            return render_template('forgot_password.html',
                                   need_request=True,
                                   user_id=user['id'],
                                   user_name=user['name'],
                                   emp_id=key)

    return render_template('forgot_password.html')


# ── 비밀번호 재설정 관리자 요청 접수 ─────────────────────────────────────────
@app.route('/forgot-password/request', methods=['POST'])
def submit_reset_request():
    user_id = request.form.get('user_id', type=int)
    emp_id  = request.form.get('emp_id', '').strip()
    if not user_id:
        flash('잘못된 요청입니다.', 'error')
        return redirect(url_for('forgot_password'))
    conn = get_db()
    user = conn.execute('SELECT name FROM users WHERE id=?', (user_id,)).fetchone()
    # 이미 대기 중인 요청이 있으면 중복 생성 방지
    existing = conn.execute(
        "SELECT id FROM password_reset_requests WHERE user_id=? AND status='대기중'",
        (user_id,)
    ).fetchone()
    if not existing:
        conn.execute(
            "INSERT INTO password_reset_requests (user_id) VALUES (?)", (user_id,)
        )
        conn.commit()
    conn.close()
    # 요청 후에도 emp_id를 유지해 인증번호 입력 링크 정상 작동
    return render_template('forgot_password.html',
                           need_request=True,
                           request_sent=True,
                           user_id=user_id,
                           user_name=user['name'] if user else '',
                           emp_id=emp_id)


# ── 관리자: 비밀번호 재설정 요청 승인 ────────────────────────────────────────
@app.route('/admin/reset-request/approve/<int:req_id>', methods=['POST'])
@admin_required
def approve_reset_request(req_id):
    import random
    conn = get_db()
    req = conn.execute(
        'SELECT r.*, u.name, u.employee_id, u.email FROM password_reset_requests r JOIN users u ON r.user_id=u.id WHERE r.id=?',
        (req_id,)
    ).fetchone()
    if not req:
        flash('요청을 찾을 수 없습니다.', 'error')
        conn.close()
        return redirect(url_for('admin'))

    from datetime import timedelta
    code    = f'{random.randint(0, 999999):06d}'
    key     = req['employee_id']
    expires = (datetime.now() + timedelta(minutes=30)).strftime('%Y-%m-%d %H:%M:%S')

    # 메모리 + DB 양쪽에 저장 (앱 재시작 후에도 유지)
    _clean_expired_codes()
    with _reset_lock:
        _reset_store[key] = {
            'code':    code,
            'user_id': req['user_id'],
            'expires': datetime.now() + timedelta(minutes=30),
        }
    conn.execute(
        "UPDATE password_reset_requests SET status='승인완료', reset_code=?, reset_expires=? WHERE id=?",
        (code, expires, req_id)
    )
    conn.commit()
    conn.close()
    flash(f'[{req["name"]}] 승인 완료 — 인증번호: {code}  (30분 유효, 사용자에게 전달하세요)', 'success')
    return redirect(url_for('admin'))


# ── 관리자: 비밀번호 재설정 요청 거절 ────────────────────────────────────────
@app.route('/admin/reset-request/reject/<int:req_id>', methods=['POST'])
@admin_required
def reject_reset_request(req_id):
    conn = get_db()
    conn.execute("UPDATE password_reset_requests SET status='거절' WHERE id=?", (req_id,))
    conn.commit()
    conn.close()
    flash('요청을 거절했습니다.', 'info')
    return redirect(url_for('admin'))


# ── 비밀번호 찾기 2단계: 인증번호 확인 ───────────────────────────────────────
@app.route('/verify-reset-code', methods=['GET', 'POST'])
def verify_reset_code():
    emp_id = request.args.get('emp_id', '') or request.form.get('emp_id', '')
    if request.method == 'POST':
        code = request.form.get('code', '').strip()
        _clean_expired_codes()
        with _reset_lock:
            entry = _reset_store.get(emp_id)

        # 메모리에 없으면 DB에서 조회 (Render 재시작 대비)
        if not entry:
            conn = get_db()
            now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            db_req = conn.execute(
                "SELECT r.user_id, r.reset_code, r.reset_expires "
                "FROM password_reset_requests r "
                "JOIN users u ON r.user_id = u.id "
                "WHERE u.employee_id=? AND r.status='승인완료' "
                "AND r.reset_code IS NOT NULL AND r.reset_code != '' "
                "AND r.reset_expires > ?",
                (emp_id, now_str)
            ).fetchone()
            conn.close()
            if db_req:
                entry = {'code': db_req['reset_code'], 'user_id': db_req['user_id']}

        if not entry:
            flash('인증번호가 만료됐거나 존재하지 않습니다. 다시 시도하세요.', 'error')
            return redirect(url_for('forgot_password'))

        if entry['code'] != code:
            flash('인증번호가 일치하지 않습니다.', 'error')
            return render_template('verify_reset_code.html', emp_id=emp_id)

        # 인증 성공 → 세션에 임시 저장 후 재설정 페이지로
        session['_reset_user_id'] = entry['user_id']
        with _reset_lock:
            _reset_store.pop(emp_id, None)
        # DB 코드도 무효화
        conn = get_db()
        conn.execute(
            "UPDATE password_reset_requests SET reset_code=NULL, reset_expires=NULL "
            "WHERE user_id=? AND status='승인완료'",
            (entry['user_id'],)
        )
        conn.commit()
        conn.close()
        return redirect(url_for('reset_password'))

    return render_template('verify_reset_code.html', emp_id=emp_id)


# ── 비밀번호 찾기 3단계: 새 비밀번호 설정 ────────────────────────────────────
@app.route('/reset-password', methods=['GET', 'POST'])
def reset_password():
    user_id = session.get('_reset_user_id')
    if not user_id:
        flash('비밀번호 재설정 세션이 없습니다. 처음부터 다시 시도하세요.', 'error')
        return redirect(url_for('forgot_password'))

    if request.method == 'POST':
        pw1 = request.form.get('password', '')
        pw2 = request.form.get('password_confirm', '')
        if len(pw1) < 6:
            flash('비밀번호는 6자 이상이어야 합니다.', 'error')
            return render_template('reset_password.html')
        if pw1 != pw2:
            flash('비밀번호가 일치하지 않습니다.', 'error')
            return render_template('reset_password.html')

        conn = get_db()
        conn.execute('UPDATE users SET password=? WHERE id=?', (hash_pw(pw1), user_id))
        conn.commit()
        conn.close()

        session.pop('_reset_user_id', None)
        session.pop('_reset_email',   None)
        flash('비밀번호가 재설정되었습니다. 새 비밀번호로 로그인하세요.', 'success')
        return redirect(url_for('login'))

    return render_template('reset_password.html')


# ── 회원가입 ──────────────────────────────────────────────────────────────────
@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        name   = request.form['name'].strip()
        emp_id = request.form['employee_id'].strip()
        email  = request.form['email'].strip()
        phone  = request.form['phone'].strip()
        team   = request.form['team'].strip()
        pw     = request.form['password']
        pw_cfm = request.form['password_confirm']
        if pw != pw_cfm:
            flash('비밀번호가 일치하지 않습니다.', 'error')
            return render_template('register.html', teams=TEAMS)
        conn = get_db()
        try:
            conn.execute(
                'INSERT INTO users (name, employee_id, email, phone, team, password) VALUES (?,?,?,?,?,?)',
                (name, emp_id, email, phone, team, hash_pw(pw))
            )
            conn.commit()
            flash('회원가입 신청이 완료되었습니다. 관리자 승인 후 로그인 가능합니다.', 'success')
            return redirect(url_for('login'))
        except Exception:
            flash('이미 사용 중인 사번입니다.', 'error')
        finally:
            conn.close()
    return render_template('register.html', teams=TEAMS)


# ── 관리자: 회원 승인 ─────────────────────────────────────────────────────────
@app.route('/admin')
@admin_required
def admin():
    conn = get_db()
    pending  = conn.execute('SELECT * FROM users WHERE is_approved=0 ORDER BY created_at DESC').fetchall()
    approved = conn.execute('SELECT * FROM users WHERE is_approved=1 AND is_admin=0 ORDER BY role, created_at DESC').fetchall()
    reset_requests = conn.execute('''
        SELECT r.id, r.status, r.created_at,
               u.name, u.employee_id, u.email
        FROM password_reset_requests r
        JOIN users u ON r.user_id = u.id
        WHERE r.status = '대기중'
        ORDER BY r.created_at DESC
    ''').fetchall()
    conn.close()
    return render_template('admin.html', pending=pending, approved=approved,
                           reset_requests=reset_requests)


# ── 데이터 관리 페이지 (관리자 전용) ──────────────────────────────────────────
@app.route('/admin/data')
@admin_required
def admin_data():
    now_ym = now_kst().strftime('%Y-%m')
    email_enabled = get_setting('email_enabled', '1') == '1'
    smtp_ok = email_config.ENABLED
    return render_template('admin_data.html',
                           now_ym=now_ym,
                           email_enabled=email_enabled,
                           smtp_ok=smtp_ok)



@app.route('/admin/approve/<int:user_id>', methods=['POST'])
@admin_required
def approve(user_id):
    role = request.form.get('role', '점검자')
    if role not in ('점검자', '승인자'):
        role = '점검자'
    conn = get_db()
    conn.execute('UPDATE users SET is_approved=1, role=? WHERE id=?', (role, user_id))
    conn.commit()
    conn.close()
    flash(f'{role}로 승인되었습니다.', 'success')
    return redirect(url_for('admin'))


@app.route('/admin/reject/<int:user_id>')
@admin_required
def reject(user_id):
    conn = get_db()
    conn.execute('DELETE FROM users WHERE id=?', (user_id,))
    conn.commit()
    conn.close()
    flash('신청이 거부되었습니다.', 'info')
    return redirect(url_for('admin'))


@app.route('/admin/change-role/<int:user_id>', methods=['POST'])
@admin_required
def change_role(user_id):
    role = request.form.get('role', '점검자')
    conn = get_db()
    conn.execute('UPDATE users SET role=? WHERE id=?', (role, user_id))
    conn.commit()
    conn.close()
    flash('역할이 변경되었습니다.', 'success')
    return redirect(url_for('admin'))


@app.route('/admin/delete-user/<int:user_id>')
@admin_required
def admin_delete_user(user_id):
    # 자기 자신은 삭제 불가
    if user_id == session.get('user_id'):
        flash('자기 자신은 삭제할 수 없습니다.', 'error')
        return redirect(url_for('admin'))
    conn = get_db()
    target = conn.execute('SELECT name FROM users WHERE id=?', (user_id,)).fetchone()
    if not target:
        conn.close()
        flash('해당 사용자를 찾을 수 없습니다.', 'error')
        return redirect(url_for('admin'))
    conn.execute('DELETE FROM users WHERE id=?', (user_id,))
    conn.commit()
    conn.close()
    flash(f'{target["name"]} 님의 계정이 삭제되었습니다.', 'success')
    return redirect(url_for('admin'))


# ── 관리자: 설비 관리 ─────────────────────────────────────────────────────────
@app.route('/admin/equipment')
@admin_required
def admin_equipment():
    q            = request.args.get('q',    '').strip()
    current_dept = request.args.get('dept', '전체').strip()
    conn = get_db()
    ph   = '%s' if conn._pg else '?'

    # 시리얼 번호 = id 기준 등록순 rank
    serial_expr = '(SELECT COUNT(*) FROM equipment e2 WHERE e2.id <= e.id)'

    wheres, params = [], []
    if q:
        wheres.append(f"(e.name LIKE {ph} OR e.location LIKE {ph} OR e.department LIKE {ph})")
        params += [f'%{q}%', f'%{q}%', f'%{q}%']
    if current_dept and current_dept != '전체':
        wheres.append(f"e.department = {ph}")
        params.append(current_dept)

    where_sql = ('WHERE ' + ' AND '.join(wheres)) if wheres else ''

    base_sql = f'''
        SELECT e.*,
               {serial_expr}             AS serial_no,
               u.name  AS creator_name,
               a.name  AS approver_name,
               COUNT(DISTINCT SUBSTR(i.inspected_at, 1, 10)) AS inspection_count,
               t.id       AS template_id,
               t.filename AS template_file
        FROM equipment e
        LEFT JOIN users u ON e.created_by  = u.id
        LEFT JOIN users a ON e.approver_id = a.id
        LEFT JOIN inspections i ON e.id = i.equipment_id
        LEFT JOIN inspection_templates t ON e.id = t.equipment_id
        {where_sql}
        GROUP BY e.id, u.name, a.name, t.id, t.filename ORDER BY e.id DESC
    '''

    equipments = conn.execute(base_sql, params).fetchall()
    approvers = conn.execute(
        "SELECT id, name, team FROM users WHERE (role='승인자' OR is_admin=1) AND is_approved=1 ORDER BY name"
    ).fetchall()
    conn.close()
    return render_template('admin_equipment.html', equipments=equipments,
                           approvers=approvers, has_qrcode=HAS_QRCODE,
                           has_openpyxl=HAS_OPENPYXL, q=q,
                           current_dept=current_dept, all_teams=TEAMS)


@app.route('/admin/equipment/set-approver/<int:eq_id>', methods=['POST'])
@admin_required
def set_equipment_approver(eq_id):
    approver_id = request.form.get('approver_id') or None
    conn = get_db()
    conn.execute('UPDATE equipment SET approver_id=? WHERE id=?', (approver_id, eq_id))
    conn.commit()
    conn.close()
    flash('승인자가 지정되었습니다.', 'success')
    return redirect(url_for('admin_equipment'))


@app.route('/admin/equipment/add', methods=['GET', 'POST'])
@admin_required
def admin_equipment_add():
    conn = get_db()
    approvers = conn.execute(
        "SELECT id, name, team FROM users WHERE (role='승인자' OR is_admin=1) AND is_approved=1 ORDER BY name"
    ).fetchall()
    all_equipments = conn.execute(
        'SELECT id, name, department FROM equipment ORDER BY name'
    ).fetchall()
    conn.close()
    if request.method == 'POST':
        name              = request.form['name'].strip()
        location          = request.form['location'].strip()
        department        = request.form['department'].strip()
        description       = request.form.get('description', '').strip()
        approver_id       = request.form.get('approver_id') or None
        qr_code           = request.form.get('qr_code', '').strip() or str(uuid.uuid4())
        mgmt_no           = request.form.get('mgmt_no', '').strip()
        manager_primary   = request.form.get('manager_primary', '').strip()
        manager_secondary = request.form.get('manager_secondary', '').strip()
        conn = get_db()
        try:
            inspection_cycle = request.form.get('inspection_cycle', '매일')
            eq_id_new = conn.insert(
                '''INSERT INTO equipment
                   (name, qr_code, location, department, description, approver_id, created_by, inspection_cycle,
                    mgmt_no, manager_primary, manager_secondary)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?)''',
                (name, qr_code, location, department, description, approver_id, session['user_id'], inspection_cycle,
                 mgmt_no, manager_primary, manager_secondary)
            )
            # 점검 항목 저장
            item_names      = request.form.getlist('item_name')
            item_categories = request.form.getlist('item_category')
            item_criterias  = request.form.getlist('item_criteria')
            item_units      = request.form.getlist('item_unit')
            item_types      = request.form.getlist('item_type')
            item_mins       = request.form.getlist('item_min')
            item_centers    = request.form.getlist('item_center')
            item_maxs       = request.form.getlist('item_max')
            for i, iname in enumerate(item_names):
                if iname.strip():
                    cat    = item_categories[i] if i < len(item_categories) else ''
                    cri    = item_criterias[i]  if i < len(item_criterias)  else ''
                    unt    = item_units[i]       if i < len(item_units)      else ''
                    itype  = item_types[i]       if i < len(item_types)      else '일반'
                    imin   = item_mins[i]        if i < len(item_mins)       else ''
                    icen   = item_centers[i]     if i < len(item_centers)    else ''
                    imax   = item_maxs[i]        if i < len(item_maxs)       else ''
                    conn.execute(
                        'INSERT INTO inspection_items (equipment_id, item_order, category, item_name, criteria, unit, item_type, min_val, center_val, max_val) VALUES (?,?,?,?,?,?,?,?,?,?)',
                        (eq_id_new, i+1, cat.strip(), iname.strip(), cri.strip(), unt.strip(),
                         itype.strip() or '일반', imin.strip(), icen.strip(), imax.strip())
                    )
            conn.commit()
            flash(f'설비 "{name}" 이(가) 등록되었습니다.', 'success')
            return redirect(url_for('admin_equipment'))
        except Exception:
            flash('이미 등록된 QR 코드입니다.', 'error')
        finally:
            conn.close()
    return render_template('admin_equipment_add.html', teams=TEAMS, approvers=approvers,
                           all_equipments=all_equipments)


@app.route('/admin/equipment/upload-template/<int:eq_id>', methods=['POST'])
@admin_required
def upload_template(eq_id):
    if not HAS_OPENPYXL:
        flash('openpyxl 패키지 필요: pip install openpyxl', 'error')
        return redirect(url_for('admin_equipment'))
    file = request.files.get('excel_file')
    if not file or not file.filename:
        flash('파일을 선택해주세요.', 'error')
        return redirect(url_for('admin_equipment'))
    if not file.filename.lower().endswith(('.xlsx', '.xls', '.xlsm')):
        flash('Excel 파일(.xlsx)만 업로드 가능합니다.', 'error')
        return redirect(url_for('admin_equipment'))
    try:
        rows_data, max_cols = parse_excel(file)
    except Exception as e:
        flash(f'파일 읽기 실패: {e}', 'error')
        return redirect(url_for('admin_equipment'))
    if not rows_data:
        flash('Excel에서 데이터를 찾을 수 없습니다.', 'error')
        return redirect(url_for('admin_equipment'))

    conn = get_db()
    if USE_PG:
        conn.execute('''
            INSERT INTO inspection_templates (equipment_id, filename, max_cols, rows)
            VALUES (?, ?, ?, ?)
            ON CONFLICT (equipment_id) DO UPDATE SET
                filename = EXCLUDED.filename,
                max_cols = EXCLUDED.max_cols,
                rows     = EXCLUDED.rows
        ''', (eq_id, file.filename, max_cols, json.dumps(rows_data, ensure_ascii=False)))
    else:
        conn.execute('''
            INSERT OR REPLACE INTO inspection_templates
                (equipment_id, filename, max_cols, rows)
            VALUES (?, ?, ?, ?)
        ''', (eq_id, file.filename, max_cols, json.dumps(rows_data, ensure_ascii=False)))
    conn.commit()
    conn.close()
    item_count = sum(1 for r in rows_data if r['is_item'])
    flash(f'점검표가 등록되었습니다. (점검 항목 {item_count}개 감지)', 'success')
    return redirect(url_for('admin_equipment'))


@app.route('/admin/equipment/delete-template/<int:eq_id>')
@admin_required
def delete_template(eq_id):
    conn = get_db()
    conn.execute('DELETE FROM inspection_templates WHERE equipment_id=?', (eq_id,))
    conn.commit()
    conn.close()
    flash('점검표가 삭제되었습니다.', 'info')
    return redirect(url_for('admin_equipment'))


@app.route('/admin/equipment/items-json/<int:eq_id>')
@admin_required
def equipment_items_json(eq_id):
    """설비 점검항목을 JSON으로 반환 (항목 복사 기능용)"""
    conn = get_db()
    eq = conn.execute('SELECT name FROM equipment WHERE id=?', (eq_id,)).fetchone()
    items = conn.execute(
        'SELECT * FROM inspection_items WHERE equipment_id=? ORDER BY item_order', (eq_id,)
    ).fetchall()
    conn.close()
    if not eq:
        return json.dumps({'error': '설비를 찾을 수 없습니다.'}), 404
    result = [dict(item) for item in items]
    return json.dumps({'eq_name': eq['name'], 'items': result}, ensure_ascii=False)


@app.route('/admin/equipment/edit/<int:eq_id>', methods=['GET', 'POST'])
@admin_required
def admin_equipment_edit(eq_id):
    conn = get_db()
    eq = conn.execute('SELECT * FROM equipment WHERE id=?', (eq_id,)).fetchone()
    if not eq:
        conn.close()
        flash('설비를 찾을 수 없습니다.', 'error')
        return redirect(url_for('admin_equipment'))
    approvers = conn.execute(
        "SELECT id, name, team FROM users WHERE (role='승인자' OR is_admin=1) AND is_approved=1 ORDER BY name"
    ).fetchall()
    # 항목 복사용: 현재 설비 제외한 전체 설비 목록
    all_equipments = conn.execute(
        'SELECT id, name, department FROM equipment WHERE id!=? ORDER BY name', (eq_id,)
    ).fetchall()
    if request.method == 'POST':
        name              = request.form['name'].strip()
        location          = request.form['location'].strip()
        department        = request.form['department'].strip()
        description       = request.form.get('description', '').strip()
        approver_id       = request.form.get('approver_id') or None
        inspection_cycle  = request.form.get('inspection_cycle', '매일')
        mgmt_no           = request.form.get('mgmt_no', '').strip()
        manager_primary   = request.form.get('manager_primary', '').strip()
        manager_secondary = request.form.get('manager_secondary', '').strip()
        conn.execute(
            '''UPDATE equipment SET name=?, location=?, department=?, description=?,
               approver_id=?, inspection_cycle=?, mgmt_no=?, manager_primary=?, manager_secondary=?
               WHERE id=?''',
            (name, location, department, description, approver_id, inspection_cycle,
             mgmt_no, manager_primary, manager_secondary, eq_id)
        )
        # ── 점검 항목 업데이트: 기존 삭제 후 새로 저장 ──
        conn.execute('DELETE FROM inspection_items WHERE equipment_id=?', (eq_id,))
        item_names      = request.form.getlist('item_name')
        item_categories = request.form.getlist('item_category')
        item_criterias  = request.form.getlist('item_criteria')
        item_units      = request.form.getlist('item_unit')
        item_types      = request.form.getlist('item_type')
        item_mins       = request.form.getlist('item_min')
        item_centers    = request.form.getlist('item_center')
        item_maxs       = request.form.getlist('item_max')
        for i, iname in enumerate(item_names):
            if iname.strip():
                cat   = item_categories[i] if i < len(item_categories) else ''
                cri   = item_criterias[i]  if i < len(item_criterias)  else ''
                unt   = item_units[i]       if i < len(item_units)      else ''
                itype = item_types[i]       if i < len(item_types)      else '일반'
                imin  = item_mins[i]        if i < len(item_mins)       else ''
                icen  = item_centers[i]     if i < len(item_centers)    else ''
                imax  = item_maxs[i]        if i < len(item_maxs)       else ''
                conn.execute(
                    'INSERT INTO inspection_items (equipment_id, item_order, category, item_name, criteria, unit, item_type, min_val, center_val, max_val) VALUES (?,?,?,?,?,?,?,?,?,?)',
                    (eq_id, i+1, cat.strip(), iname.strip(), cri.strip(), unt.strip(),
                     itype.strip() or '일반', imin.strip(), icen.strip(), imax.strip())
                )
        conn.commit()
        conn.close()
        flash(f'설비 "{name}" 정보 및 점검항목이 수정되었습니다.', 'success')
        return redirect(url_for('admin_equipment'))
    db_items = conn.execute(
        'SELECT * FROM inspection_items WHERE equipment_id=? ORDER BY item_order', (eq_id,)
    ).fetchall()
    conn.close()
    items_json = json.dumps([dict(item) for item in db_items], ensure_ascii=False)
    return render_template('admin_equipment_edit.html', eq=eq, teams=TEAMS, approvers=approvers,
                           db_items=db_items, items_json=items_json,
                           all_equipments=all_equipments)


@app.route('/admin/equipment/delete/<int:eq_id>')
@admin_required
def admin_equipment_delete(eq_id):
    conn = get_db()
    conn.execute('DELETE FROM equipment WHERE id=?', (eq_id,))
    conn.commit()
    conn.close()
    flash('설비가 삭제되었습니다.', 'info')
    return redirect(url_for('admin_equipment'))


def _make_qr_label(eq_name, qr_url, serial_no):
    """QR 코드 + 설비명 + 시리얼 번호 라벨 이미지 생성"""
    # QR 코드 생성
    qr = qrcode.QRCode(version=1, box_size=10, border=3)
    qr.add_data(qr_url)
    qr.make(fit=True)
    qr_img = qr.make_image(fill_color='black', back_color='white').convert('RGB')
    qr_w, qr_h = qr_img.size

    # 라벨 캔버스 크기 (위 여백 50 + QR + 아래 여백 90)
    pad_top    = 50
    pad_bottom = 90
    label_w = qr_w
    label_h = pad_top + qr_h + pad_bottom

    canvas = Image.new('RGB', (label_w, label_h), 'white')
    canvas.paste(qr_img, (0, pad_top))
    draw = ImageDraw.Draw(canvas)

    # 한글 지원 폰트 우선 탐색 (Windows / Linux Render 모두 지원)
    _base = os.path.dirname(os.path.abspath(__file__))
    _FONT_CANDIDATES = [
        os.path.join(_base, 'static', 'fonts', 'malgunbd.ttf'),  # 프로젝트 내장 폰트 (최우선)
        "C:/Windows/Fonts/malgunbd.ttf",
        "C:/Windows/Fonts/malgun.ttf",
        "C:/Windows/Fonts/gulim.ttc",
        "/usr/share/fonts/truetype/nanum/NanumGothicBold.ttf",
        "/usr/share/fonts/truetype/nanum/NanumGothic.ttf",
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
    ]
    font_name = font_sn = None
    for _fp in _FONT_CANDIDATES:
        try:
            font_name = ImageFont.truetype(_fp, 22)
            font_sn   = ImageFont.truetype(_fp, 28)
            break
        except Exception:
            continue
    if font_name is None:
        font_name = ImageFont.load_default()
        font_sn   = font_name

    # 설비명 (상단)
    bbox = draw.textbbox((0, 0), eq_name, font=font_name)
    tw = bbox[2] - bbox[0]
    draw.text(((label_w - tw) // 2, 10), eq_name, fill='black', font=font_name)

    # 시리얼 넘버 (하단)
    sn_text = f"S/N : {serial_no}"
    bbox2 = draw.textbbox((0, 0), sn_text, font=font_sn)
    tw2 = bbox2[2] - bbox2[0]
    draw.text(((label_w - tw2) // 2, pad_top + qr_h + 10), sn_text, fill='#f97316', font=font_sn)

    # 구분선
    draw.line([(20, pad_top + qr_h + 6), (label_w - 20, pad_top + qr_h + 6)], fill='#e5e7eb', width=1)

    buf = BytesIO()
    canvas.save(buf, format='PNG')
    buf.seek(0)
    return buf


@app.route('/admin/equipment/qr/<int:eq_id>')
@admin_required
def equipment_qr_download(eq_id):
    if not HAS_QRCODE:
        flash('QR 생성 패키지 필요: pip install "qrcode[pil]"', 'error')
        return redirect(url_for('admin_equipment'))
    conn = get_db()
    eq = conn.execute('SELECT * FROM equipment WHERE id=?', (eq_id,)).fetchone()
    # 등록순 시리얼 번호: id 기준 오름차순 rank
    rank = conn.execute(
        'SELECT COUNT(*) AS cnt FROM equipment WHERE id <= ?', (eq_id,)
    ).fetchone()['cnt']
    conn.close()
    if not eq:
        return '설비를 찾을 수 없습니다.', 404

    serial_no = f'{rank:04d}'
    qr_url    = request.host_url.rstrip('/') + url_for('qr_redirect', code=eq['qr_code'])
    buf       = _make_qr_label(eq['name'], qr_url, serial_no)
    return send_file(buf, mimetype='image/png', as_attachment=True,
                     download_name=f'{eq["name"]}_{serial_no}_QR라벨.png')


# ── QR 전체 인쇄 ─────────────────────────────────────────────────────────────
@app.route('/admin/equipment/qr-print')
@admin_required
def equipment_qr_print():
    if not HAS_QRCODE:
        flash('QR 생성 패키지 필요: pip install "qrcode[pil]"', 'error')
        return redirect(url_for('admin_equipment'))

    # 팀 필터
    selected_dept = request.args.get('dept', '').strip()

    conn = get_db()
    ph = '%s' if conn._pg else '?'
    if selected_dept and selected_dept != '전체':
        equipments = conn.execute(
            f'SELECT id, name, qr_code, location, department, mgmt_no '
            f'FROM equipment WHERE department = {ph} ORDER BY name',
            (selected_dept,)
        ).fetchall()
    else:
        equipments = conn.execute(
            'SELECT id, name, qr_code, location, department, mgmt_no FROM equipment ORDER BY name'
        ).fetchall()
    conn.close()

    host_url = request.host_url.rstrip('/')
    qr_items = []
    for i, eq in enumerate(equipments):
        qr_url = host_url + url_for('qr_redirect', code=eq['qr_code'])
        qr_obj = qrcode.QRCode(version=1, box_size=6, border=2)
        qr_obj.add_data(qr_url)
        qr_obj.make(fit=True)
        qr_img = qr_obj.make_image(fill_color='black', back_color='white')
        buf = BytesIO()
        qr_img.save(buf, format='PNG')
        b64 = base64.b64encode(buf.getvalue()).decode('utf-8')
        qr_items.append({
            'name':     eq['name'],
            'location': eq['location'] or '',
            'dept':     eq['department'] or '',
            'mgmt_no':  eq['mgmt_no']  or '',
            'serial':   f'{i + 1:04d}',
            'qr_b64':   b64,
        })

    # 35개(5×7)씩 페이지 분할
    pages = [qr_items[i:i + 35] for i in range(0, len(qr_items), 35)]
    dept_label = selected_dept if (selected_dept and selected_dept != '전체') else '전체'
    return render_template('qr_print.html',
                           pages=pages,
                           total=len(qr_items),
                           all_teams=TEAMS,
                           selected_dept=dept_label)


# ── QR 리다이렉트 ─────────────────────────────────────────────────────────────
@app.route('/qr/<code>')
def qr_redirect(code):
    conn = get_db()
    eq = conn.execute('SELECT * FROM equipment WHERE qr_code=?', (code,)).fetchone()
    conn.close()
    if not eq:
        flash('등록되지 않은 QR 코드입니다.', 'error')
        return redirect(url_for('login'))
    if not session.get('user_id'):
        return redirect(url_for('login', next=url_for('inspect', eq_id=eq['id'])))
    return redirect(url_for('inspect', eq_id=eq['id']))


# ── 설비 점검 페이지 ──────────────────────────────────────────────────────────
@app.route('/inspect/<int:eq_id>', methods=['GET', 'POST'])
@login_required
def inspect(eq_id):
    conn = get_db()
    eq = conn.execute('''
        SELECT e.*, a.name AS approver_name
        FROM equipment e
        LEFT JOIN users a ON e.approver_id = a.id
        WHERE e.id=?
    ''', (eq_id,)).fetchone()

    if not eq:
        conn.close()
        flash('설비를 찾을 수 없습니다.', 'error')
        return redirect(url_for('dashboard'))

    is_approver = (
        session.get('is_admin') or
        (session.get('role') == '승인자' and eq['approver_id'] == session['user_id'])
    )
    is_inspector = session.get('role') == '점검자' or session.get('is_admin')

    tmpl = conn.execute(
        'SELECT * FROM inspection_templates WHERE equipment_id=?', (eq_id,)
    ).fetchone()
    tmpl_rows     = json.loads(tmpl['rows']) if tmpl else None
    tmpl_max_cols = tmpl['max_cols']         if tmpl else 0

    db_items = conn.execute(
        'SELECT * FROM inspection_items WHERE equipment_id=? ORDER BY item_order',
        (eq_id,)
    ).fetchall()

    # 오늘 이미 점검됐는지 확인
    today_insp = conn.execute(f'''
        SELECT i.*, u.name AS inspector_name
        FROM inspections i
        JOIN users u ON i.inspector_id = u.id
        WHERE i.equipment_id=? AND {conn.date_col("i.inspected_at")}={conn.today}
        AND i.status IN ('점검완료','승인완료')
        ORDER BY i.inspected_at DESC LIMIT 1
    ''', (eq_id,)).fetchone()

    # 오늘 점검 세부항목 조회
    today_insp_details = []
    if today_insp:
        today_insp_details = conn.execute('''
            SELECT d.row_index, d.result, d.detail_notes, d.item_id,
                   ii.item_name, ii.category, ii.criteria, ii.unit,
                   ii.item_type, ii.min_val, ii.center_val, ii.max_val
            FROM inspection_details d
            LEFT JOIN inspection_items ii ON d.item_id = ii.id
            WHERE d.inspection_id = ?
            ORDER BY d.row_index
        ''', (today_insp['id'],)).fetchall()

    if request.method == 'POST':
        action = request.form.get('action')

        if action == 'submit' and is_inspector:
            # ── 점검 날짜 처리 ──────────────────────────────────────
            today_str = now_kst().strftime('%Y-%m-%d')
            inspect_date = request.form.get('inspect_date', '').strip()
            if not inspect_date:
                inspect_date = today_str
            try:
                datetime.strptime(inspect_date, '%Y-%m-%d')
            except ValueError:
                inspect_date = today_str
            # 미래 날짜 방지
            if inspect_date > today_str:
                inspect_date = today_str

            # 선택 날짜 중복 점검 방지
            date_insp = conn.execute(f'''
                SELECT id FROM inspections
                WHERE equipment_id=? AND {conn.date_col("inspected_at")}=?
                AND status IN ('점검완료','승인완료')
            ''', (eq_id, inspect_date)).fetchone()
            if date_insp:
                flash(f'{inspect_date} 날짜에 이미 점검이 완료되었습니다. 중복 점검은 불가합니다.', 'warning')
                return redirect(url_for('inspect', eq_id=eq_id))

            # 저장할 inspected_at: 선택 날짜 + 현재 시각(KST)
            inspected_at = inspect_date + ' ' + now_kst().strftime('%H:%M:%S')

            overall_notes = request.form.get('notes', '').strip()

            if db_items:
                item_results = []
                for item in db_items:
                    itype = (item.get('item_type') or '일반')
                    iid   = item['id']
                    if itype == '수치':
                        # 수리중/휴동 선택 여부 확인
                        special = request.form.get(f'special_item_{iid}', '')
                        if special in ('수리중', '휴동'):
                            r_val = special
                            n_val = request.form.get(f'notes_item_{iid}', '')
                        else:
                            # 수치 입력값 처리
                            numeric_str = request.form.get(f'numeric_val_{iid}', '').strip()
                            unit_label  = item.get('unit', '') or ''
                            n_val = f"{numeric_str} {unit_label}".strip() if numeric_str else ''
                            if numeric_str:
                                try:
                                    num   = float(numeric_str)
                                    min_s = (item.get('min_val') or '').strip()
                                    max_s = (item.get('max_val') or '').strip()
                                    in_range = True
                                    if min_s:
                                        try:
                                            if num < float(min_s): in_range = False
                                        except ValueError:
                                            pass
                                    if max_s:
                                        try:
                                            if num > float(max_s): in_range = False
                                        except ValueError:
                                            pass
                                    r_val = '정상' if in_range else '이상'
                                except (ValueError, TypeError):
                                    r_val = '이상'
                            else:
                                r_val = '정상'  # 값 미입력 = 정상
                    else:
                        r_val = request.form.get(f'result_item_{iid}', '정상')
                        n_val = request.form.get(f'notes_item_{iid}', '')
                    item_results.append((iid, r_val, n_val))
                all_vals = [r for _, r, _ in item_results]
                overall  = ('이상' if '이상' in all_vals else
                            '수리중' if '수리중' in all_vals else
                            '휴동' if all(r in ('휴동','해당없음') for r in all_vals) else '정상')
                ins_id = conn.insert(
                    "INSERT INTO inspections (equipment_id, inspector_id, result, notes, status, inspected_at) VALUES (?,?,?,?,'점검완료',?)",
                    (eq_id, session['user_id'], overall, overall_notes, inspected_at)
                )
                for item_id, r_val, n_val in item_results:
                    conn.execute(
                        'INSERT INTO inspection_details (inspection_id, row_index, item_id, result, detail_notes) VALUES (?,?,?,?,?)',
                        (ins_id, 0, item_id, r_val, n_val)
                    )
                conn.commit()
                result = overall

            elif tmpl_rows:
                item_results = []
                for idx, row in enumerate(tmpl_rows):
                    if not row['is_item']:
                        continue
                    r_val = request.form.get(f'result_{idx}', '정상')
                    n_val = request.form.get(f'notes_{idx}', '')
                    item_results.append((idx, r_val, n_val))

                all_vals = [r for _, r, _ in item_results]
                overall  = ('이상' if '이상' in all_vals else
                            '수리중' if '수리중' in all_vals else
                            '휴동' if all(r in ('휴동','해당없음') for r in all_vals) else '정상')

                ins_id = conn.insert(
                    '''INSERT INTO inspections
                       (equipment_id, inspector_id, result, notes, status, inspected_at)
                       VALUES (?,?,?,?,'점검완료',?)''',
                    (eq_id, session['user_id'], overall, overall_notes, inspected_at)
                )
                for idx, r_val, n_val in item_results:
                    conn.execute(
                        '''INSERT INTO inspection_details
                           (inspection_id, row_index, result, detail_notes)
                           VALUES (?,?,?,?)''',
                        (ins_id, idx, r_val, n_val)
                    )
                conn.commit()
                result = overall
            else:
                result = request.form.get('result', '정상')
                conn.execute(
                    '''INSERT INTO inspections
                       (equipment_id, inspector_id, result, notes, status, inspected_at)
                       VALUES (?,?,?,?,'점검완료',?)''',
                    (eq_id, session['user_id'], result, overall_notes, inspected_at)
                )
                conn.commit()

            # ── 주1회/월1회 자동 채움 ─────────────────────────────────────────
            cycle = eq.get('inspection_cycle', '일1회') or '일1회'
            if cycle in ('주1회', '월1회'):
                insp_date_str = inspected_at[:10]
                auto_cnt = _auto_fill_cycle(
                    conn, eq_id, session['user_id'], eq.get('approver_id'),
                    result, insp_date_str, cycle
                )
                conn.commit()
                if auto_cnt > 0:
                    flash(f'({cycle}) 나머지 {auto_cnt}일 자동 점검완료 처리됐습니다.', 'info')

            if eq['approver_id']:
                approver = conn.execute(
                    'SELECT name, email FROM users WHERE id=?', (eq['approver_id'],)
                ).fetchone()
                if approver and approver['email']:
                    send_approval_request(
                        to_email       = approver['email'],
                        approver_name  = approver['name'],
                        inspector_name = session['user_name'],
                        eq_name        = eq['name'],
                        location       = eq['location'] or '-',
                        result         = result,
                        notes          = overall_notes,
                        eq_id          = eq_id,
                        host_url       = request.host_url,
                    )

            flash('점검이 완료되었습니다. 승인자에게 알림이 발송됩니다.', 'success')
            return redirect(url_for('inspect', eq_id=eq_id))

        elif action == 'resubmit' and is_inspector:
            # ── 승인 대기 중 점검 수정 제출 ───────────────────────────
            if not today_insp or today_insp['status'] != '점검완료':
                flash('수정할 수 있는 점검 기록이 없습니다. (이미 승인됐거나 기록 없음)', 'warning')
                return redirect(url_for('inspect', eq_id=eq_id))

            ins_id        = today_insp['id']
            inspected_at  = today_insp['inspected_at']  # 원래 날짜·시각 유지
            overall_notes = request.form.get('notes', '').strip()

            if db_items:
                item_results = []
                for item in db_items:
                    itype = item['item_type'] or '일반'
                    iid   = item['id']
                    if itype == '수치':
                        special = request.form.get(f'special_item_{iid}', '')
                        if special in ('수리중', '휴동'):
                            r_val = special
                            n_val = request.form.get(f'notes_item_{iid}', '')
                        else:
                            numeric_str = request.form.get(f'numeric_val_{iid}', '').strip()
                            unit_label  = item['unit'] or ''
                            n_val = f"{numeric_str} {unit_label}".strip() if numeric_str else ''
                            if numeric_str:
                                try:
                                    num   = float(numeric_str)
                                    min_s = (item['min_val'] or '').strip()
                                    max_s = (item['max_val'] or '').strip()
                                    in_range = True
                                    if min_s:
                                        try:
                                            if num < float(min_s): in_range = False
                                        except ValueError: pass
                                    if max_s:
                                        try:
                                            if num > float(max_s): in_range = False
                                        except ValueError: pass
                                    r_val = '정상' if in_range else '이상'
                                except (ValueError, TypeError):
                                    r_val = '이상'
                            else:
                                r_val = '정상'
                    else:
                        r_val = request.form.get(f'result_item_{iid}', '정상')
                        n_val = request.form.get(f'notes_item_{iid}', '')
                    item_results.append((iid, r_val, n_val))

                all_vals = [r for _, r, _ in item_results]
                overall  = ('이상'  if '이상'  in all_vals else
                            '수리중' if '수리중' in all_vals else
                            '휴동'  if all(r in ('휴동','해당없음') for r in all_vals) else '정상')

                # 기존 세부항목 삭제 후 재저장
                conn.execute('DELETE FROM inspection_details WHERE inspection_id=?', (ins_id,))
                conn.execute(
                    'UPDATE inspections SET result=?, notes=?, inspected_at=? WHERE id=?',
                    (overall, overall_notes, inspected_at, ins_id)
                )
                for item_id, r_val, n_val in item_results:
                    conn.execute(
                        'INSERT INTO inspection_details (inspection_id, row_index, item_id, result, detail_notes) VALUES (?,?,?,?,?)',
                        (ins_id, 0, item_id, r_val, n_val)
                    )
                conn.commit()

            elif tmpl_rows:
                item_results = []
                for idx, row in enumerate(tmpl_rows):
                    if not row['is_item']:
                        continue
                    r_val = request.form.get(f'result_{idx}', '정상')
                    n_val = request.form.get(f'notes_{idx}', '')
                    item_results.append((idx, r_val, n_val))
                all_vals = [r for _, r, _ in item_results]
                overall  = ('이상'  if '이상'  in all_vals else
                            '수리중' if '수리중' in all_vals else
                            '휴동'  if all(r in ('휴동','해당없음') for r in all_vals) else '정상')
                conn.execute('DELETE FROM inspection_details WHERE inspection_id=?', (ins_id,))
                conn.execute(
                    'UPDATE inspections SET result=?, notes=?, inspected_at=? WHERE id=?',
                    (overall, overall_notes, inspected_at, ins_id)
                )
                for idx, r_val, n_val in item_results:
                    conn.execute(
                        'INSERT INTO inspection_details (inspection_id, row_index, result, detail_notes) VALUES (?,?,?,?)',
                        (ins_id, idx, r_val, n_val)
                    )
                conn.commit()

            else:
                result = request.form.get('result', '정상')
                conn.execute(
                    'UPDATE inspections SET result=?, notes=?, inspected_at=? WHERE id=?',
                    (result, overall_notes, inspected_at, ins_id)
                )
                conn.commit()

            flash('점검 내용이 수정되었습니다.', 'success')
            return redirect(url_for('inspect', eq_id=eq_id))

        elif action == 'approve' and is_approver:
            ins_id = request.form.get('inspection_id')
            approved_ins = conn.execute(
                'SELECT inspected_at FROM inspections WHERE id=?', (ins_id,)
            ).fetchone()
            conn.execute(
                f'''UPDATE inspections
                   SET status='승인완료', approved_by=?, approved_at={conn.now_fn}
                   WHERE id=? AND status='점검완료' ''',
                (session['user_id'], ins_id)
            )
            # 같은 날 같은 설비의 나머지 대기 건 모두 삭제 (실제 점검 날짜 기준)
            if approved_ins:
                ins_date = approved_ins['inspected_at'][:10]
                conn.execute(
                    f'''DELETE FROM inspections
                        WHERE equipment_id=? AND status='점검완료' AND id!=?
                        AND {conn.date_col("inspected_at")}=?''',
                    (eq_id, ins_id, ins_date)
                )
            conn.commit()
            flash('승인이 완료되었습니다.', 'success')
            return redirect(url_for('inspect', eq_id=eq_id))

    pending_approvals = []
    if is_approver:
        pending_approvals = conn.execute('''
            SELECT i.*, u.name AS inspector_name
            FROM inspections i
            JOIN users u ON i.inspector_id = u.id
            WHERE i.equipment_id=? AND i.status='점검완료'
            ORDER BY i.inspected_at DESC
        ''', (eq_id,)).fetchall()

    history = conn.execute(f'''
        SELECT i.*, u.name AS inspector_name, a.name AS approved_name
        FROM inspections i
        JOIN users u ON i.inspector_id = u.id
        LEFT JOIN users a ON i.approved_by = a.id
        WHERE i.equipment_id=?
          AND i.id = (
              SELECT MAX(id) FROM inspections
              WHERE equipment_id = i.equipment_id
                AND {conn.date_col("inspected_at")} = {conn.date_col("i.inspected_at")}
          )
        ORDER BY i.inspected_at DESC
        LIMIT 20
    ''', (eq_id,)).fetchall()
    conn.close()

    now = now_kst()
    today_date = now.strftime('%Y-%m-%d')

    # 수정용 기존 점검 데이터 JSON (승인 대기 중일 때 폼 사전 입력용)
    details_for_edit = []
    if today_insp and today_insp['status'] == '점검완료':
        for d in today_insp_details:
            details_for_edit.append({
                'item_id':      d['item_id'],
                'result':       d['result'] or '',
                'detail_notes': d['detail_notes'] or '',
                'item_type':    d['item_type'] or '일반',
            })
    details_json = json.dumps(details_for_edit, ensure_ascii=False)

    return render_template('inspect.html', eq=eq, history=history,
                           pending_approvals=pending_approvals,
                           is_approver=is_approver, is_inspector=is_inspector,
                           tmpl_rows=tmpl_rows, tmpl_max_cols=tmpl_max_cols,
                           db_items=db_items, today_insp=today_insp,
                           today_insp_details=today_insp_details,
                           today_date=today_date,
                           now_year=now.year, now_month=now.month,
                           details_json=details_json)


# ── 점검 결과 수정 ───────────────────────────────────────────────────────────
@app.route('/inspection/<int:ins_id>/edit', methods=['GET', 'POST'])
@login_required
def inspection_edit(ins_id):
    conn = get_db()
    insp = conn.execute('''
        SELECT i.*, e.name AS eq_name, e.id AS eq_id, e.approver_id,
               u.name AS inspector_name
        FROM inspections i
        JOIN equipment e ON i.equipment_id = e.id
        JOIN users u ON i.inspector_id = u.id
        WHERE i.id = ?
    ''', (ins_id,)).fetchone()

    if not insp:
        conn.close()
        flash('점검 기록을 찾을 수 없습니다.', 'error')
        return redirect(request.referrer or url_for('dashboard'))

    is_admin     = session.get('is_admin')
    is_inspector = insp['inspector_id'] == session['user_id']
    is_approver  = (session.get('role') == '승인자' and
                    insp['approver_id'] == session['user_id'])

    # 권한: 관리자는 항상 / 점검자는 미승인일 때 / 승인자는 미승인일 때
    can_edit = is_admin or (
        (is_inspector or is_approver) and insp['status'] == '점검완료'
    )
    if not can_edit:
        conn.close()
        flash('수정 권한이 없거나 이미 승인된 점검입니다.', 'error')
        return redirect(request.referrer or url_for('dashboard'))

    db_items = conn.execute(
        'SELECT * FROM inspection_items WHERE equipment_id=? ORDER BY item_order',
        (insp['eq_id'],)
    ).fetchall()

    details = {}
    if db_items:
        for d in conn.execute(
            'SELECT * FROM inspection_details WHERE inspection_id=?', (ins_id,)
        ).fetchall():
            details[d['item_id']] = d

    if request.method == 'POST':
        new_result = request.form.get('result', insp['result'])
        new_notes  = request.form.get('notes', '').strip()

        if db_items:
            item_results = []
            for item in db_items:
                r_val = request.form.get(f'result_item_{item["id"]}', '정상')
                n_val = request.form.get(f'notes_item_{item["id"]}', '')
                item_results.append((item['id'], r_val, n_val))
            all_vals = [r for _, r, _ in item_results]
            new_result = ('이상'   if '이상'   in all_vals else
                          '수리중' if '수리중' in all_vals else
                          '수리필요' if '수리필요' in all_vals else
                          '휴동'   if all(r in ('휴동', '해당없음') for r in all_vals) else '정상')
            for item_id, r_val, n_val in item_results:
                conn.execute('''
                    UPDATE inspection_details SET result=?, detail_notes=?
                    WHERE inspection_id=? AND item_id=?
                ''', (r_val, n_val, ins_id, item_id))

        conn.execute(
            'UPDATE inspections SET result=?, notes=? WHERE id=?',
            (new_result, new_notes, ins_id)
        )
        conn.commit()
        conn.close()
        flash('점검 결과가 수정되었습니다.', 'success')
        return redirect(request.referrer or url_for('my_inspections'))

    conn.close()
    return render_template('inspection_edit.html',
                           insp=insp, db_items=db_items, details=details)


# ── 일별 점검 결과 (전체 설비) ───────────────────────────────────────────────
@app.route('/approve-inspection/<int:ins_id>', methods=['POST'])
@login_required
def approve_inspection(ins_id):
    conn = get_db()
    insp = conn.execute(
        'SELECT i.*, e.id AS eq_id FROM inspections i JOIN equipment e ON i.equipment_id=e.id WHERE i.id=?',
        (ins_id,)
    ).fetchone()
    if not insp:
        flash('점검 기록을 찾을 수 없습니다.', 'error')
        conn.close()
        return redirect(url_for('dashboard'))

    # 승인자 권한 확인
    if insp['equipment_id'] and not (session.get('is_admin') or
       conn.execute('SELECT id FROM equipment WHERE id=? AND approver_id=?',
                    (insp['equipment_id'], session['user_id'])).fetchone()):
        flash('승인 권한이 없습니다.', 'error')
        conn.close()
        return redirect(url_for('dashboard'))

    conn.execute(
        f"UPDATE inspections SET status='승인완료', approved_by=?, approved_at={conn.now_fn} WHERE id=?",
        (session['user_id'], ins_id)
    )
    # 같은 날 같은 설비의 나머지 대기 건 모두 삭제 (실제 점검 날짜 기준)
    ins_date = insp['inspected_at'][:10]
    conn.execute(
        f'''DELETE FROM inspections
            WHERE equipment_id=? AND status='점검완료' AND id!=?
            AND {conn.date_col("inspected_at")}=?''',
        (insp['equipment_id'], ins_id, ins_date)
    )
    conn.commit()
    conn.close()
    flash('승인이 완료되었습니다.', 'success')
    return redirect(request.referrer or url_for('dashboard'))


# ── 점검 초기화 (관리자 전용) ─────────────────────────────────────────────────
@app.route('/admin/reset-inspection', methods=['POST'])
@login_required
def reset_inspection():
    if not session.get('is_admin'):
        flash('관리자만 초기화할 수 있습니다.', 'error')
        return redirect(request.referrer or url_for('dashboard'))

    equipment_id = request.form.get('equipment_id', type=int)
    date_str     = request.form.get('date', '')          # 'YYYY-MM-DD'

    if not equipment_id or not date_str:
        flash('잘못된 요청입니다.', 'error')
        return redirect(request.referrer or url_for('dashboard'))

    conn = get_db()
    eq   = conn.execute('SELECT name FROM equipment WHERE id=?', (equipment_id,)).fetchone()
    if not eq:
        conn.close()
        flash('설비를 찾을 수 없습니다.', 'error')
        return redirect(request.referrer or url_for('dashboard'))

    # 해당 날짜의 점검 ID 목록 조회 (세부항목 삭제용)
    ins_ids = conn.execute(
        f"SELECT id FROM inspections WHERE equipment_id=? AND {conn.date_col('inspected_at')}=?",
        (equipment_id, date_str)
    ).fetchall()

    for row in ins_ids:
        conn.execute('DELETE FROM inspection_details WHERE inspection_id=?', (row['id'],))

    deleted = conn.execute(
        f"DELETE FROM inspections WHERE equipment_id=? AND {conn.date_col('inspected_at')}=?",
        (equipment_id, date_str)
    )
    conn.commit()
    conn.close()

    flash(f'[{eq["name"]}] {date_str} 점검 기록이 초기화되었습니다.', 'success')
    return redirect(request.referrer or url_for('daily_results'))


# ── 월별 점검 기록 전체 삭제 GET 버전 (긴급용, 관리자 전용) ──────────────────
@app.route('/admin/delete-month/<ym>')
@admin_required
def admin_delete_month_get(ym):
    """URL 직접 접근용: /admin/delete-month/2026-05"""
    if not ym or len(ym) != 7:
        flash('올바른 연월 형식이 아닙니다 (예: 2026-05).', 'error')
        return redirect(url_for('admin'))
    conn = get_db()
    if conn._pg:
        result = conn.execute(
            "DELETE FROM inspections WHERE LEFT(inspected_at,7) = %s", (ym,))
    else:
        result = conn.execute(
            "DELETE FROM inspections WHERE substr(inspected_at,1,7) = ?", (ym,))
    cnt = result.rowcount if hasattr(result, 'rowcount') else '?'
    conn.commit()
    conn.close()
    flash(f'✅ {ym} 점검 기록 {cnt}건이 삭제되었습니다.', 'success')
    return redirect(url_for('admin_data'))


# ── 월별 점검 기록 전체 삭제 (관리자 전용) ────────────────────────────────────
@app.route('/admin/delete-month', methods=['POST'])
@admin_required
def admin_delete_month():
    ym = request.form.get('ym', '').strip()   # 형식: 2026-05
    if not ym or len(ym) != 7:
        flash('올바른 연월을 입력하세요 (예: 2026-05).', 'error')
        return redirect(url_for('admin_data'))
    conn = get_db()
    if conn._pg:
        result = conn.execute(
            "DELETE FROM inspections WHERE LEFT(inspected_at,7) = %s", (ym,))
    else:
        result = conn.execute(
            "DELETE FROM inspections WHERE substr(inspected_at,1,7) = ?", (ym,))
    cnt = result.rowcount if hasattr(result, 'rowcount') else '?'
    conn.commit()
    conn.close()
    flash(f'{ym} 점검 기록 {cnt}건이 삭제되었습니다.', 'success')
    return redirect(url_for('admin_data'))


# ── 이메일 알림 ON/OFF 토글 (관리자 전용) ────────────────────────────────────
@app.route('/admin/toggle-email', methods=['POST'])
@admin_required
def admin_toggle_email():
    current = get_setting('email_enabled', '1')
    new_val = '0' if current == '1' else '1'
    set_setting('email_enabled', new_val)
    state = 'ON' if new_val == '1' else 'OFF'
    flash(f'이메일 알림이 {state}으로 변경되었습니다.', 'success')
    return redirect(url_for('admin_data'))


# ── 일별 점검 기록 삭제 (관리자 전용) ────────────────────────────────────────
@app.route('/admin/delete-day', methods=['POST'])
@admin_required
def admin_delete_day():
    date_str = request.form.get('del_date', '').strip()   # 형식: 2026-05-16
    if not date_str or len(date_str) != 10:
        flash('올바른 날짜를 입력하세요 (예: 2026-05-16).', 'error')
        return redirect(url_for('admin_data'))
    conn = get_db()
    if conn._pg:
        result = conn.execute(
            "DELETE FROM inspections WHERE LEFT(inspected_at,10) = %s", (date_str,))
    else:
        result = conn.execute(
            "DELETE FROM inspections WHERE substr(inspected_at,1,10) = ?", (date_str,))
    cnt = result.rowcount if hasattr(result, 'rowcount') else '?'
    conn.commit()
    conn.close()
    flash(f'{date_str} 점검 기록 {cnt}건이 삭제되었습니다.', 'success')
    return redirect(url_for('admin_data'))


# ── 날짜별 전 설비 휴동 처리 (관리자 전용) ───────────────────────────────────
@app.route('/admin/bulk-idle', methods=['POST'])
@admin_required
def admin_bulk_idle():
    """지정한 날짜의 모든 설비에 대해 기존 기록을 삭제하고 휴동 기록을 삽입한다."""
    date_str = request.form.get('idle_date', '').strip()   # 형식: 2026-05-01
    if not date_str or len(date_str) != 10:
        flash('올바른 날짜를 입력하세요 (예: 2026-05-01).', 'error')
        return redirect(url_for('admin'))

    conn = get_db()
    # 관리자 계정 id (기록용 점검자로 사용)
    admin_id = session['user_id']
    inspected_ts = date_str + ' 00:00:00'

    equipments = conn.execute('SELECT id FROM equipment').fetchall()

    inserted = 0
    for eq in equipments:
        eq_id = eq['id']
        # 해당 날짜 기존 기록 모두 삭제
        if conn._pg:
            conn.execute(
                "DELETE FROM inspections WHERE equipment_id=%s AND LEFT(inspected_at,10)=%s",
                (eq_id, date_str))
        else:
            conn.execute(
                "DELETE FROM inspections WHERE equipment_id=? AND substr(inspected_at,1,10)=?",
                (eq_id, date_str))

        # 승인자 조회
        eq_row = conn.execute('SELECT approver_id FROM equipment WHERE id=?'
                              if not conn._pg else
                              'SELECT approver_id FROM equipment WHERE id=%s',
                              (eq_id,)).fetchone()
        approver_id = eq_row['approver_id'] if eq_row else None

        if conn._pg:
            conn.execute(
                """INSERT INTO inspections
                   (equipment_id, inspector_id, result, status, notes,
                    inspected_at, approved_by, approved_at)
                   VALUES (%s,%s,'휴동','승인완료','공휴일/휴동',%s,%s,NOW())""",
                (eq_id, admin_id, inspected_ts, approver_id or admin_id))
        else:
            conn.execute(
                """INSERT INTO inspections
                   (equipment_id, inspector_id, result, status, notes,
                    inspected_at, approved_by, approved_at)
                   VALUES (?,?,'휴동','승인완료','공휴일/휴동',?,?,datetime('now','localtime'))""",
                (eq_id, admin_id, inspected_ts, approver_id or admin_id))
        inserted += 1

    conn.commit()
    conn.close()
    flash(f'{date_str} — {inserted}개 설비 휴동 처리 완료되었습니다.', 'success')
    return redirect(url_for('admin_data'))


# ── 대시보드: 오늘 날짜 전 설비 휴동 처리 (관리자 전용 빠른 버튼) ──────────────
@app.route('/dashboard/bulk-idle-today', methods=['POST'])
@admin_required
def dashboard_bulk_idle_today():
    """오늘 날짜의 모든 설비를 즉시 휴동 처리하고 대시보드로 복귀."""
    today_str = now_kst().strftime('%Y-%m-%d')
    admin_id  = session['user_id']
    conn      = get_db()
    ph        = '%s' if conn._pg else '?'

    equipments = conn.execute('SELECT id, approver_id FROM equipment').fetchall()
    for eq in equipments:
        eq_id       = eq['id']
        approver_id = eq['approver_id'] or admin_id
        ts          = today_str + ' 00:00:00'
        if conn._pg:
            conn.execute("DELETE FROM inspections WHERE equipment_id=%s AND LEFT(inspected_at,10)=%s",
                         (eq_id, today_str))
            conn.execute("""INSERT INTO inspections
                (equipment_id, inspector_id, result, status, notes, inspected_at, approved_by, approved_at)
                VALUES (%s,%s,'휴동','승인완료','공휴일/휴동',%s,%s,NOW())""",
                (eq_id, admin_id, ts, approver_id))
        else:
            conn.execute("DELETE FROM inspections WHERE equipment_id=? AND substr(inspected_at,1,10)=?",
                         (eq_id, today_str))
            conn.execute("""INSERT INTO inspections
                (equipment_id, inspector_id, result, status, notes, inspected_at, approved_by, approved_at)
                VALUES (?,?,'휴동','승인완료','공휴일/휴동',?,?,datetime('now','localtime'))""",
                (eq_id, admin_id, ts, approver_id))

    conn.commit()
    conn.close()
    flash(f'{today_str} — {len(equipments)}개 설비 휴동 처리 완료되었습니다.', 'success')
    return redirect(url_for('dashboard'))


# ── 중복 점검 기록 일괄 정리 (관리자 전용) ────────────────────────────────────
@app.route('/admin/cleanup-duplicates', methods=['POST'])
@login_required
def cleanup_duplicates():
    if not session.get('is_admin'):
        flash('관리자만 실행할 수 있습니다.', 'error')
        return redirect(url_for('dashboard'))

    conn = get_db()
    # 날짜·설비별로 MAX(id)만 남기고 나머지 삭제
    if conn._pg:
        result = conn.execute(f'''
            DELETE FROM inspections
            WHERE id NOT IN (
                SELECT MAX(id) FROM inspections
                GROUP BY equipment_id, {conn.date_col("inspected_at")}
            )
        ''')
    else:
        result = conn.execute(f'''
            DELETE FROM inspections
            WHERE id NOT IN (
                SELECT MAX(id) FROM inspections
                GROUP BY equipment_id, {conn.date_col("inspected_at")}
            )
        ''')
    conn.commit()
    conn.close()
    flash('중복 점검 기록 정리가 완료되었습니다.', 'success')
    return redirect(url_for('admin_data'))


@app.route('/bulk-inspect', methods=['GET', 'POST'])
@login_required
def bulk_inspect():
    import traceback as _tb
    try:
        return _bulk_inspect_inner()
    except Exception as _outer_e:
        _detail = _tb.format_exc()
        print(f'[bulk_inspect OUTER ERROR] {_outer_e}\n{_detail}', flush=True)
        try:
            app.logger.error(f'[bulk_inspect OUTER ERROR] {_outer_e}\n{_detail}')
        except Exception:
            pass
        flash(f'일괄점검 오류: {_outer_e}', 'error')
        return redirect(url_for('dashboard'))


def _bulk_inspect_inner():
    conn = get_db()
    today_str = now_kst().strftime('%Y-%m-%d')

    # DB Row → 순수 Python dict 변환 (POST/GET 공용)
    def _to_dict(row):
        if row is None:
            return None
        if isinstance(row, dict):
            return dict(row)
        return {k: row[k] for k in row.keys()}

    if request.method == 'POST':
        import traceback as _tb
        import time as _time
        _t0 = _time.time()
        print(f'[bulk_inspect POST] 시작 v2-batch', flush=True)

        inspect_date = request.form.get('inspect_date', '').strip()
        if not inspect_date:
            inspect_date = today_str
        try:
            datetime.strptime(inspect_date, '%Y-%m-%d')
        except ValueError:
            inspect_date = today_str
        if inspect_date > today_str:
            inspect_date = today_str

        inspected_at = inspect_date + ' ' + now_kst().strftime('%H:%M:%S')
        eq_id_strs   = request.form.getlist('eq_ids')
        eq_ids_int   = []
        for s in eq_id_strs:
            try:
                eq_ids_int.append(int(s))
            except (ValueError, TypeError):
                pass

        if not eq_ids_int:
            conn.close()
            flash('처리할 설비가 없습니다.', 'warning')
            return redirect(url_for('bulk_inspect'))

        print(f'[bulk_inspect POST] eq_ids={len(eq_ids_int)}개, pg={conn._pg}', flush=True)

        # ── ① 사전 일괄 조회 (DB 왕복 최소화) ─────────────────────────────────
        ph = ','.join(['%s' if conn._pg else '?' for _ in eq_ids_int])

        # 이미 점검된 설비 ID 집합
        done_ids = set(
            r['equipment_id'] for r in conn.execute(
                f"SELECT DISTINCT equipment_id FROM inspections "
                f"WHERE equipment_id IN ({ph}) "
                f"AND {conn.date_col('inspected_at')}=? "
                f"AND status IN ('점검완료','승인완료')",
                eq_ids_int + [inspect_date]
            ).fetchall()
        )

        # 설비 정보 맵
        eq_map = {
            r['id']: _to_dict(r)
            for r in conn.execute(
                f"SELECT * FROM equipment WHERE id IN ({ph})",
                eq_ids_int
            ).fetchall()
        }

        # 점검 항목 맵
        items_map = {}
        for r in conn.execute(
            f"SELECT * FROM inspection_items WHERE equipment_id IN ({ph}) ORDER BY item_order",
            eq_ids_int
        ).fetchall():
            items_map.setdefault(r['equipment_id'], []).append(_to_dict(r))

        # 승인자 정보 맵
        approver_ids = list({eq['approver_id'] for eq in eq_map.values()
                             if eq.get('approver_id')})
        approver_map = {}
        if approver_ids:
            aph = ','.join(['%s' if conn._pg else '?' for _ in approver_ids])
            for r in conn.execute(
                f"SELECT id, name, email FROM users WHERE id IN ({aph})",
                approver_ids
            ).fetchall():
                approver_map[r['id']] = _to_dict(r)

        # ── ② 폼 데이터를 순수 Python으로 처리 (DB 없음) ──────────────────────
        skip_count  = 0
        email_tasks = []

        # 저장할 데이터: [(eq_id, overall_result, notes, [(item_id,r,n), ...])]
        to_save = []

        for eq_id in eq_ids_int:
            if request.form.get(f'skip_eq_{eq_id}'):
                skip_count += 1
                continue
            if eq_id in done_ids:
                skip_count += 1
                continue

            eq = eq_map.get(eq_id)
            if not eq:
                continue

            db_items      = items_map.get(eq_id, [])
            overall_notes = request.form.get(f'notes_eq_{eq_id}', '').strip()

            if db_items:
                item_results = []
                for item in db_items:
                    itype = item.get('item_type') or '일반'
                    iid   = item['id']
                    if itype == '수치':
                        special = request.form.get(f'special_item_{iid}', '')
                        if special in ('수리중', '휴동'):
                            r_val = special
                            n_val = request.form.get(f'notes_item_{iid}', '')
                        else:
                            numeric_str = request.form.get(f'numeric_val_{iid}', '').strip()
                            unit_label  = item.get('unit') or ''
                            n_val = f"{numeric_str} {unit_label}".strip() if numeric_str else ''
                            if numeric_str:
                                try:
                                    num   = float(numeric_str)
                                    min_s = (item.get('min_val') or '').strip()
                                    max_s = (item.get('max_val') or '').strip()
                                    in_range = True
                                    if min_s:
                                        try:
                                            if num < float(min_s): in_range = False
                                        except ValueError:
                                            pass
                                    if max_s:
                                        try:
                                            if num > float(max_s): in_range = False
                                        except ValueError:
                                            pass
                                    r_val = '정상' if in_range else '이상'
                                except (ValueError, TypeError):
                                    r_val = '이상'
                            else:
                                r_val = '정상'
                    else:
                        r_val = request.form.get(f'result_item_{iid}', '정상')
                        n_val = request.form.get(f'notes_item_{iid}', '')
                    item_results.append((iid, r_val, n_val))

                all_vals = [r for _, r, _ in item_results]
                overall  = ('이상'   if '이상'   in all_vals else
                            '수리중' if '수리중' in all_vals else
                            '휴동'   if all(r in ('휴동', '해당없음') for r in all_vals) else '정상')
            else:
                overall      = request.form.get(f'simple_result_eq_{eq_id}', '정상')
                item_results = []

            to_save.append((eq_id, overall, overall_notes, item_results))

            # 이메일 큐
            approver_id = eq.get('approver_id')
            if approver_id:
                approver = approver_map.get(approver_id)
                if approver and approver.get('email'):
                    email_tasks.append(dict(
                        to_email       = approver['email'],
                        approver_name  = approver['name'],
                        inspector_name = session['user_name'],
                        eq_name        = eq.get('name', ''),
                        location       = eq.get('location') or '-',
                        result         = overall,
                        notes          = overall_notes,
                        eq_id          = eq_id,
                        host_url       = request.host_url,
                    ))

        success_count = 0
        fail_names    = []

        # ── ③ 배치 INSERT: inspections 전체를 한 번에 (RETURNING id) ────────
        try:
            if not to_save:
                conn.close()
                flash(f'일괄 점검 완료 ✅  0건 처리 / {skip_count}건 건너뜀', 'success')
                return redirect(url_for('bulk_inspect'))

            uid = session['user_id']
            print(f'[bulk_inspect POST] ③ 배치 INSERT 시작: {len(to_save)}건, elapsed={_time.time()-_t0:.2f}s', flush=True)

            if conn._pg:
                # PostgreSQL: VALUES (...),(...),... RETURNING id  → 1회 왕복
                vals_sql = ','.join(['(%s,%s,%s,%s,%s,%s)'] * len(to_save))
                params   = []
                for eq_id, overall, overall_notes, _ in to_save:
                    params.extend([eq_id, uid, overall, overall_notes, '점검완료', inspected_at])
                cur = conn._conn.cursor()
                cur.execute(
                    f"INSERT INTO inspections "
                    f"(equipment_id,inspector_id,result,notes,status,inspected_at) "
                    f"VALUES {vals_sql} RETURNING id",
                    params
                )
                ins_ids = [row[0] for row in cur.fetchall()]
            else:
                # SQLite: executemany 후 rowid 직접 조회
                ins_ids = []
                raw = conn._conn
                for eq_id, overall, overall_notes, _ in to_save:
                    cur2 = raw.execute(
                        "INSERT INTO inspections "
                        "(equipment_id,inspector_id,result,notes,status,inspected_at) "
                        "VALUES (?,?,?,?,'점검완료',?)",
                        (eq_id, uid, overall, overall_notes, inspected_at)
                    )
                    ins_ids.append(cur2.lastrowid)

            success_count = len(ins_ids)
            print(f'[bulk_inspect POST] ③ inspections 완료: {success_count}건, elapsed={_time.time()-_t0:.2f}s', flush=True)

            # ── ④ 배치 INSERT: inspection_details 전체를 한 번에 ─────────────
            details_rows = []
            for i, (eq_id, overall, overall_notes, item_results) in enumerate(to_save):
                ins_id = ins_ids[i]
                for item_id, r_val, n_val in item_results:
                    details_rows.append((ins_id, 0, item_id, r_val, n_val))

            if details_rows:
                if conn._pg:
                    vals_sql2 = ','.join(['(%s,%s,%s,%s,%s)'] * len(details_rows))
                    params2   = [v for row in details_rows for v in row]
                    cur3 = conn._conn.cursor()
                    cur3.execute(
                        f"INSERT INTO inspection_details "
                        f"(inspection_id,row_index,item_id,result,detail_notes) "
                        f"VALUES {vals_sql2}",
                        params2
                    )
                else:
                    conn._conn.executemany(
                        "INSERT INTO inspection_details "
                        "(inspection_id,row_index,item_id,result,detail_notes) "
                        "VALUES (?,?,?,?,?)",
                        details_rows
                    )

            print(f'[bulk_inspect POST] ④ details {len(details_rows)}행 준비, elapsed={_time.time()-_t0:.2f}s', flush=True)

            # ── ⑤ 1회 커밋 ──────────────────────────────────────────────────
            conn._conn.commit()
            print(f'[bulk_inspect POST] ⑤ 커밋 완료, elapsed={_time.time()-_t0:.2f}s', flush=True)

        except Exception as bulk_err:
            err_detail = _tb.format_exc()
            app.logger.error(f'[bulk_inspect] 배치 저장 실패: {bulk_err}\n{err_detail}')
            print(f'[bulk_inspect BATCH ERROR] {bulk_err}\n{err_detail}', flush=True)
            try:
                conn._conn.rollback()
            except Exception:
                pass
            success_count = 0
            fail_names = [eq_map.get(eq_id, {}).get('name', str(eq_id))
                          for eq_id, *_ in to_save]
        finally:
            try:
                conn.close()
            except Exception:
                pass

        # ── ⑥ 이메일 일괄 발송 (백그라운드 스레드) ───────────────────────────
        for task in email_tasks:
            try:
                send_approval_request(**task)
            except Exception:
                pass

        if fail_names:
            flash(f'일괄 점검 오류 ⚠️  저장 실패 ({len(fail_names)}건). '
                  f'다시 시도해 주세요.', 'error')
        else:
            flash(f'일괄 점검 완료 ✅  {success_count}건 처리 / {skip_count}건 건너뜀', 'success')
        return redirect(url_for('bulk_inspect'))

    # GET: 날짜 파라미터 (없으면 오늘)
    selected_date = request.args.get('date', today_str).strip()
    try:
        datetime.strptime(selected_date, '%Y-%m-%d')
    except ValueError:
        selected_date = today_str
    if selected_date > today_str:
        selected_date = today_str

    eq_data      = []
    current_dept = '전체'

    try:
        print(f'[bulk_inspect GET] selected_date={selected_date}', flush=True)
        dept_sql_g, dept_params_g, current_dept = _dept_filter(conn)
        equipments = conn.execute(f'''
            SELECT e.id, e.name, e.location, e.department, e.approver_id,
                   e.inspection_cycle, a.name AS approver_name
            FROM equipment e
            LEFT JOIN users a ON e.approver_id = a.id
            WHERE 1=1 {dept_sql_g}
            ORDER BY e.name
        ''', dept_params_g).fetchall()
        print(f'[bulk_inspect GET] equipments count={len(equipments)}', flush=True)

        all_eq_ids = [r['id'] for r in equipments]

        # ── 일괄 조회: 점검 항목 (1 query) ────────────────────────────────────
        items_map_get = {}
        if all_eq_ids:
            g_ph = ','.join(['%s' if conn._pg else '?' for _ in all_eq_ids])
            for r in conn.execute(
                f'SELECT * FROM inspection_items WHERE equipment_id IN ({g_ph}) ORDER BY item_order',
                all_eq_ids
            ).fetchall():
                items_map_get.setdefault(r['equipment_id'], []).append(_to_dict(r))

        # ── 일괄 조회: 이미 완료된 점검 (1 query) ────────────────────────────
        done_eq_ids_get = set()
        if all_eq_ids:
            for r in conn.execute(
                f"SELECT DISTINCT equipment_id FROM inspections "
                f"WHERE equipment_id IN ({g_ph}) "
                f"AND {conn.date_col('inspected_at')}=? "
                f"AND status IN ('점검완료','승인완료')",
                all_eq_ids + [selected_date]
            ).fetchall():
                done_eq_ids_get.add(r['equipment_id'])

        eq_data = []
        for eq_row in equipments:
            eq_id = eq_row['id']
            eq_data.append({
                'eq_dict':     _to_dict(eq_row),
                'items_list':  items_map_get.get(eq_id, []),
                'already_done': eq_id in done_eq_ids_get,
            })
        print(f'[bulk_inspect GET] eq_data built ok (3 queries), rendering template', flush=True)
    except Exception as e:
        import traceback
        err_detail = traceback.format_exc()
        app.logger.error(f'bulk_inspect GET error: {e}\n{err_detail}')
        print(f'[bulk_inspect GET ERROR] {e}\n{err_detail}', flush=True)
        try:
            conn.rollback()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass
        flash(f'페이지 로드 중 오류가 발생했습니다: {e}', 'error')
        return redirect(url_for('dashboard'))

    try:
        conn.close()
    except Exception:
        pass
    try:
        return render_template('bulk_inspect.html', eq_data=eq_data,
                               today_date=today_str, selected_date=selected_date,
                               current_dept=current_dept, all_teams=TEAMS)
    except Exception as e:
        app.logger.exception('bulk_inspect render error')
        flash(f'화면 렌더링 오류: {e}', 'error')
        return redirect(url_for('dashboard'))


# ── 일괄 승인 ─────────────────────────────────────────────────────────────────
@app.route('/bulk-approve', methods=['GET', 'POST'])
@login_required
def bulk_approve():
    # 승인자 또는 관리자만 접근 허용
    if not (session.get('is_admin') or session.get('role') == '승인자'):
        flash('승인 권한이 없습니다.', 'error')
        return redirect(url_for('dashboard'))

    conn = get_db()
    today_str = now_kst().strftime('%Y-%m-%d')

    def _to_dict(row):
        if row is None:
            return None
        if isinstance(row, dict):
            return dict(row)
        return {k: row[k] for k in row.keys()}

    if request.method == 'POST':
        selected_date = request.form.get('approve_date', today_str).strip()
        try:
            datetime.strptime(selected_date, '%Y-%m-%d')
        except ValueError:
            selected_date = today_str
        if selected_date > today_str:
            selected_date = today_str

        ins_ids = request.form.getlist('ins_ids')
        if not ins_ids:
            flash('승인할 항목을 선택하세요.', 'warning')
            conn.close()
            return redirect(url_for('bulk_approve', date=selected_date))

        success_count = 0
        try:
            for ins_id in ins_ids:
                insp = _to_dict(conn.execute(
                    "SELECT * FROM inspections WHERE id=? AND status='점검완료'", (ins_id,)
                ).fetchone())
                if not insp:
                    continue

                # 관리자가 아닌 경우 해당 설비의 승인자인지 확인
                if not session.get('is_admin'):
                    eq_check = conn.execute(
                        'SELECT id FROM equipment WHERE id=? AND approver_id=?',
                        (insp['equipment_id'], session['user_id'])
                    ).fetchone()
                    if not eq_check:
                        continue

                conn.execute(
                    f'''UPDATE inspections
                           SET status='승인완료', approved_by=?, approved_at={conn.now_fn}
                         WHERE id=? AND status='점검완료' ''',
                    (session['user_id'], ins_id)
                )
                # 같은 날 같은 설비의 나머지 대기 건 모두 삭제
                ins_date = insp['inspected_at'][:10]
                conn.execute(
                    f'''DELETE FROM inspections
                         WHERE equipment_id=? AND status='점검완료' AND id!=?
                           AND {conn.date_col("inspected_at")}=?''',
                    (insp['equipment_id'], ins_id, ins_date)
                )
                success_count += 1
        except Exception as e:
            app.logger.exception('bulk_approve POST error')
            conn.close()
            flash(f'승인 처리 중 오류가 발생했습니다: {e}', 'error')
            return redirect(url_for('bulk_approve', date=selected_date))

        conn.commit()
        conn.close()
        flash(f'일괄 승인 완료 ✅  {success_count}건 처리', 'success')
        return redirect(url_for('bulk_approve', date=selected_date))

    # GET
    selected_date = request.args.get('date', today_str).strip()
    try:
        datetime.strptime(selected_date, '%Y-%m-%d')
    except ValueError:
        selected_date = today_str
    if selected_date > today_str:
        selected_date = today_str

    try:
        if session.get('is_admin'):
            rows = conn.execute(f'''
                SELECT i.id, i.equipment_id, i.result, i.notes, i.inspected_at,
                       e.name AS eq_name, e.location, e.department,
                       u.name AS inspector_name,
                       a.name AS approver_name
                FROM inspections i
                JOIN equipment e ON i.equipment_id = e.id
                JOIN users u ON i.inspector_id = u.id
                LEFT JOIN users a ON e.approver_id = a.id
                WHERE i.status = '점검완료'
                  AND {conn.date_col("i.inspected_at")} = ?
                ORDER BY e.name, i.inspected_at
            ''', (selected_date,)).fetchall()
        else:
            rows = conn.execute(f'''
                SELECT i.id, i.equipment_id, i.result, i.notes, i.inspected_at,
                       e.name AS eq_name, e.location, e.department,
                       u.name AS inspector_name,
                       a.name AS approver_name
                FROM inspections i
                JOIN equipment e ON i.equipment_id = e.id
                JOIN users u ON i.inspector_id = u.id
                LEFT JOIN users a ON e.approver_id = a.id
                WHERE i.status = '점검완료'
                  AND e.approver_id = ?
                  AND {conn.date_col("i.inspected_at")} = ?
                ORDER BY e.name, i.inspected_at
            ''', (session['user_id'], selected_date)).fetchall()

        pending_list = [_to_dict(r) for r in rows]
    except Exception as e:
        app.logger.exception('bulk_approve GET error')
        conn.close()
        flash(f'페이지 로드 중 오류가 발생했습니다: {e}', 'error')
        return redirect(url_for('dashboard'))

    conn.close()
    return render_template('bulk_approve.html',
                           pending_list=pending_list,
                           today_date=today_str,
                           selected_date=selected_date)


@app.route('/daily-results')
@login_required
def daily_results():
    today = now_kst().strftime('%Y-%m-%d')
    selected_date = request.args.get('date', today)

    conn = get_db()
    dept_sql, dept_params, current_dept = _dept_filter(conn)
    # 설비당 해당 날짜 최선 점검 1건 (승인완료 우선, 이후 최신순)
    rows = conn.execute(f'''
        SELECT e.id AS eq_id, e.name AS eq_name, e.location, e.department,
               e.manager_primary, e.manager_secondary,
               i.id AS insp_id, i.result, i.status, i.inspected_at,
               u.name AS inspector_name, a.name AS approved_name
        FROM equipment e
        LEFT JOIN inspections i ON i.id = (
            SELECT id FROM inspections
            WHERE equipment_id = e.id
              AND {conn.date_col("inspected_at")} = ?
            ORDER BY CASE WHEN status='승인완료' THEN 0 ELSE 1 END, inspected_at DESC
            LIMIT 1
        )
        LEFT JOIN users u ON i.inspector_id = u.id
        LEFT JOIN users a ON i.approved_by = a.id
        WHERE 1=1 {dept_sql}
        ORDER BY e.name
    ''', [selected_date] + dept_params).fetchall()
    conn.close()

    return render_template('daily_results.html',
                           rows=rows, selected_date=selected_date,
                           current_dept=current_dept, all_teams=TEAMS)


# ── 내 점검 결과 ──────────────────────────────────────────────────────────────
@app.route('/my-inspections')
@login_required
def my_inspections():
    date_from     = request.args.get('date_from', '')
    date_to       = request.args.get('date_to', '')
    result_filter = request.args.get('result', '')

    conn = get_db()

    # 설비·날짜별 최선 1건 (승인완료 우선, 이후 최신순)
    date_filter = ''
    if date_from:
        date_filter += f' AND {conn.date_col("i.inspected_at")} >= ?'
    if date_to:
        date_filter += f' AND {conn.date_col("i.inspected_at")} <= ?'
    result_f = ' AND i.result = ?' if result_filter else ''

    params = [session['user_id']]
    if date_from: params.append(date_from)
    if date_to:   params.append(date_to)
    if result_filter: params.append(result_filter)

    query = f'''
        SELECT i.*,
               e.name       AS eq_name,
               e.location   AS eq_location,
               e.department AS eq_dept,
               a.name       AS approved_name
        FROM inspections i
        JOIN equipment e ON i.equipment_id = e.id
        LEFT JOIN users a ON i.approved_by = a.id
        WHERE i.inspector_id = ?
          {date_filter}{result_f}
          AND i.id = (
              SELECT MAX(id) FROM inspections
              WHERE equipment_id = i.equipment_id
                AND inspector_id = i.inspector_id
                AND {conn.date_col("inspected_at")} = {conn.date_col("i.inspected_at")}
          )
        ORDER BY i.inspected_at DESC
    '''
    records = conn.execute(query, params).fetchall()
    conn.close()

    # 중복 제거된 records 기준으로 통계 계산
    stats = {
        'total':    len(records),
        'normal':   sum(1 for r in records if r['result'] == '정상'),
        'abnormal': sum(1 for r in records if r['result'] == '이상'),
        'repair':   sum(1 for r in records if r['result'] in ('수리필요', '수리중')),
        'idle':     sum(1 for r in records if r['result'] == '휴동'),
        'approved': sum(1 for r in records if r['status'] == '승인완료'),
    }

    return render_template('my_inspections.html', records=records, stats=stats,
                           date_from=date_from, date_to=date_to,
                           result_filter=result_filter)

# ── 도움말 ────────────────────────────────────────────────────────────────────
@app.route("/help")
@login_required
def help_page():
    return render_template("help.html")


# ── 내 승인 결과 (승인자 전용) ───────────────────────────────────────────────
@app.route('/my-approvals')
@login_required
def my_approvals():
    date_from     = request.args.get('date_from', '')
    date_to       = request.args.get('date_to', '')
    result_filter = request.args.get('result', '')

    conn = get_db()

    date_filter = ''
    if date_from:
        date_filter += f' AND {conn.date_col("i.inspected_at")} >= ?'
    if date_to:
        date_filter += f' AND {conn.date_col("i.inspected_at")} <= ?'
    result_f = ' AND i.result = ?' if result_filter else ''

    params = [session['user_id']]
    if date_from: params.append(date_from)
    if date_to:   params.append(date_to)
    if result_filter: params.append(result_filter)

    records = conn.execute(f'''
        SELECT i.*,
               e.name       AS eq_name,
               e.location   AS eq_location,
               e.department AS eq_dept,
               u.name       AS inspector_name
        FROM inspections i
        JOIN equipment e ON i.equipment_id = e.id
        JOIN users u ON i.inspector_id = u.id
        WHERE i.approved_by = ? AND i.status = '승인완료'
          {date_filter}{result_f}
        ORDER BY i.approved_at DESC
    ''', params).fetchall()
    conn.close()

    stats = {
        'total':    len(records),
        'normal':   sum(1 for r in records if r['result'] == '정상'),
        'abnormal': sum(1 for r in records if r['result'] == '이상'),
        'repair':   sum(1 for r in records if r['result'] in ('수리필요', '수리중')),
        'idle':     sum(1 for r in records if r['result'] == '휴동'),
    }

    return render_template('my_approvals.html', records=records, stats=stats,
                           date_from=date_from, date_to=date_to,
                           result_filter=result_filter)


# ── 대시보드 ──────────────────────────────────────────────────────────────────
@app.route('/dashboard')
@login_required
def dashboard():
    conn = get_db()

    # 설비별 오늘 최선 점검 1건 (승인완료 우선, 이후 최신순)
    today_list = conn.execute(f'''
        SELECT i.id, i.result, i.status, i.inspected_at, i.notes,
               e.id AS equipment_id, e.name AS eq_name, e.location AS eq_location,
               a.name AS approved_name
        FROM inspections i
        JOIN equipment e ON i.equipment_id = e.id
        LEFT JOIN users a ON i.approved_by = a.id
        WHERE i.inspector_id=?
          AND {conn.date_col("i.inspected_at")}={conn.today}
          AND i.id = (
              SELECT id FROM inspections
              WHERE equipment_id = i.equipment_id
                AND {conn.date_col("inspected_at")}={conn.today}
              ORDER BY CASE WHEN status='승인완료' THEN 0 ELSE 1 END, inspected_at DESC
              LIMIT 1
          )
        ORDER BY i.inspected_at DESC
    ''', (session['user_id'],)).fetchall()

    today_count = len(today_list)

    total_eq = conn.execute(
        'SELECT COUNT(*) AS cnt FROM equipment'
    ).fetchone()['cnt']

    # 오늘 휴동 여부: 주말이거나 전 설비가 휴동 처리된 날
    _today_kst = now_kst()
    today_is_idle = _today_kst.weekday() >= 5  # 토(5)/일(6)
    if not today_is_idle and total_eq > 0:
        idle_cnt = conn.execute(f'''
            SELECT COUNT(DISTINCT equipment_id) AS cnt
            FROM inspections
            WHERE {conn.date_col("inspected_at")} = {conn.today}
              AND result = '휴동'
        ''').fetchone()['cnt']
        today_is_idle = (idle_cnt >= total_eq)

    pending_list = []
    approved_count = 0
    if session.get('role') == '승인자' or session.get('is_admin'):
        pending_list = conn.execute(f'''
            SELECT i.id, i.result, i.inspected_at,
                   e.id AS eq_id, e.name AS eq_name,
                   u.name AS inspector_name
            FROM inspections i
            JOIN equipment e ON i.equipment_id = e.id
            JOIN users u ON i.inspector_id = u.id
            WHERE e.approver_id=? AND i.status='점검완료'
              AND i.result != '휴동'
              AND i.id = (
                  SELECT MAX(id) FROM inspections
                  WHERE equipment_id = i.equipment_id
                    AND status='점검완료'
                    AND {conn.date_col("inspected_at")} = {conn.date_col("i.inspected_at")}
              )
            ORDER BY i.inspected_at DESC
        ''', (session['user_id'],)).fetchall()

        approved_count = conn.execute(f'''
            SELECT COUNT(DISTINCT i.equipment_id) AS cnt
            FROM inspections i
            JOIN equipment e ON i.equipment_id = e.id
            WHERE e.approver_id=? AND i.status='승인완료'
              AND i.result != '휴동'
              AND {conn.date_col("i.inspected_at")}={conn.today}
        ''', (session['user_id'],)).fetchone()['cnt']

    conn.close()
    return render_template('dashboard.html', today_count=today_count,
                           today_list=today_list, today_is_idle=today_is_idle,
                           today_str=_today_kst.strftime('%Y-%m-%d'),
                           total_eq=total_eq, pending_list=pending_list,
                           approved_count=approved_count)


# ── 모니터링 (점검율 대시보드) ───────────────────────────────────────────────
@app.route('/monitoring')
@login_required
def monitoring():
    import json as _json
    from collections import defaultdict as _dd
    now   = now_kst()
    year  = int(request.args.get('year',  now.year))
    month = int(request.args.get('month', now.month))
    ym    = f"{year}-{month:02d}"

    conn = get_db()

    # 팀 필터: dept_sql 은 'AND e.department = ?' 형태 (alias='e' 기본값)
    # → 아래 쿼리에서 모두 equipment 에 'e' 별칭을 붙여서 사용
    dept_sql_m, dept_params_m, current_dept = _dept_filter(conn)

    ph = '%s' if conn._pg else '?'   # 파라미터 플레이스홀더

    days_in_month = calendar.monthrange(year, month)[1]
    passed_days   = now.day if (year == now.year and month == now.month) else days_in_month
    today_str     = now.strftime('%Y-%m-%d')

    if conn._pg:
        yr_expr  = "LEFT(i.inspected_at, 4)"
        ym_expr  = "LEFT(i.inspected_at, 7)"
        day_expr = "LEFT(i.inspected_at, 10)"
    else:
        yr_expr  = "strftime('%Y', i.inspected_at)"
        ym_expr  = "strftime('%Y-%m', i.inspected_at)"
        day_expr = "date(i.inspected_at)"

    try:
        # ── 전체 설비 수 (alias e 사용) ──────────────────────────────────────
        total_eq = conn.execute(
            f'SELECT COUNT(*) AS cnt FROM equipment e WHERE 1=1{dept_sql_m}',
            dept_params_m
        ).fetchone()['cnt']

        # ── 오늘 점검 완료 수 (휴동 제외) ───────────────────────────────────────
        today_done = conn.execute(f'''
            SELECT COUNT(DISTINCT i.equipment_id) AS cnt
            FROM inspections i
            JOIN equipment e ON i.equipment_id = e.id
            WHERE {conn.date_col("i.inspected_at")} = {ph}
              AND i.status IN ('점검완료','승인완료')
              AND i.result != {ph}
              {dept_sql_m}
        ''', [today_str, '휴동'] + dept_params_m).fetchone()['cnt']

        # 오늘 휴동 여부
        today_idle_cnt = conn.execute(f'''
            SELECT COUNT(DISTINCT i.equipment_id) AS cnt
            FROM inspections i
            JOIN equipment e ON i.equipment_id = e.id
            WHERE {conn.date_col("i.inspected_at")} = {ph}
              AND i.result = {ph}
              {dept_sql_m}
        ''', [today_str, '휴동'] + dept_params_m).fetchone()['cnt']
        # 오늘이 주말(토=5, 일=6)이거나 DB 전체 휴동이면 휴동일
        today_is_idle = (
            (total_eq > 0 and today_idle_cnt >= total_eq) or
            now.weekday() >= 5
        )
        today_rate = 0 if today_is_idle else (round(today_done / total_eq * 100, 1) if total_eq else 0)

        # ── 해당 월 점검 완료 쌍 (휴동 제외) ───────────────────────────────────
        insp_pairs_raw = conn.execute(f'''
            SELECT DISTINCT i.equipment_id, {day_expr} AS day
            FROM inspections i
            JOIN equipment e ON i.equipment_id = e.id
            WHERE {ym_expr} = {ph}
              AND i.status IN ('점검완료','승인완료')
              AND i.result != {ph}
              {dept_sql_m}
        ''', [ym, '휴동'] + dept_params_m).fetchall()

        # ── 해당 월 휴동 기록 ──────────────────────────────────────────────────
        idle_pairs_raw = conn.execute(f'''
            SELECT DISTINCT i.equipment_id, {day_expr} AS day
            FROM inspections i
            JOIN equipment e ON i.equipment_id = e.id
            WHERE {ym_expr} = {ph}
              AND i.result = {ph}
              {dept_sql_m}
        ''', [ym, '휴동'] + dept_params_m).fetchall()

        # ── 전체 설비 목록 (id, department) ──────────────────────────────────
        all_eq_rows = conn.execute(
            f'SELECT e.id, e.name, e.department, e.location FROM equipment e WHERE 1=1{dept_sql_m} ORDER BY e.name',
            dept_params_m
        ).fetchall()

        # ── 해당 연도 전체 점검 완료 쌍 (연간 월별 집계용, 휴동 제외) ────────────
        yearly_insp_raw = conn.execute(f'''
            SELECT DISTINCT i.equipment_id, {day_expr} AS day
            FROM inspections i
            JOIN equipment e ON i.equipment_id = e.id
            WHERE {yr_expr} = {ph}
              AND i.status IN ('점검완료','승인완료')
              AND i.result != {ph}
              {dept_sql_m}
        ''', [str(year), '휴동'] + dept_params_m).fetchall()

    except Exception as _e:
        conn.close()
        import traceback as _tb
        print(f'[monitoring ERROR] {_e}\n{_tb.format_exc()}', flush=True)
        flash(f'모니터링 조회 오류: {_e}', 'error')
        return redirect(url_for('dashboard'))

    conn.close()

    # ── Python 단 집계 ────────────────────────────────────────────────────────
    # eq_id → department 빠른 조회용 dict
    eq_dept_map = {r['id']: (r['department'] or '미지정') for r in all_eq_rows}

    # ── ① 휴동일 먼저 확정 (DB 기록 + 주말 자동) ─────────────────────────────
    idle_day_eq = _dd(set)
    for r in idle_pairs_raw:
        day_str2 = str(r['day'])
        day_num  = int(day_str2.split('-')[2])
        idle_day_eq[day_num].add(r['equipment_id'])

    idle_days_set = set()
    for d in range(1, passed_days + 1):
        # DB에 전체 설비 휴동 기록이 있는 날
        if total_eq and len(idle_day_eq[d]) >= total_eq:
            idle_days_set.add(d)
        # 주말(토=5, 일=6) 자동 휴동
        elif datetime(year, month, d).weekday() >= 5:
            idle_days_set.add(d)

    passed_days_work = max(passed_days - len(idle_days_set), 0)

    # ── ② 점검 집계 (휴동일 제외) ────────────────────────────────────────────
    # 일별: day_num → set of equipment_ids
    day_eq_set = _dd(set)
    # 팀별: dept → set of (eq_id, day_str) 중복제거
    dept_done  = _dd(set)
    # 설비별: eq_id → 점검일수
    eq_done_days = _dd(int)

    for r in insp_pairs_raw:
        day_str2 = str(r['day'])
        day_num  = int(day_str2.split('-')[2])
        if day_num in idle_days_set:
            continue           # 휴동일 점검은 집계 제외
        eq_id   = r['equipment_id']
        dept_nm = eq_dept_map.get(eq_id, '미지정')

        day_eq_set[day_num].add(eq_id)
        dept_done[dept_nm].add((eq_id, day_str2))
        eq_done_days[eq_id] += 1

    # 일별 점검율 (차트) — 휴동일은 None(null)으로
    chart_labels = [f"{month}/{d}" for d in range(1, passed_days + 1)]
    chart_values = []
    for d in range(1, passed_days + 1):
        if d in idle_days_set:
            chart_values.append(None)
        else:
            chart_values.append(round(len(day_eq_set[d]) / total_eq * 100, 1) if total_eq else 0)
    non_idle_vals = [v for v in chart_values if v is not None]
    avg_rate = round(sum(non_idle_vals) / len(non_idle_vals), 1) if non_idle_vals else 0

    # 팀별 집계
    dept_eq = _dd(set)
    for r in all_eq_rows:
        dept_eq[r['department'] or '미지정'].add(r['id'])

    dept_data = []
    for dept_nm, eq_ids in sorted(dept_eq.items()):
        eq_cnt       = len(eq_ids)
        done_cnt     = len(dept_done[dept_nm])
        max_possible = eq_cnt * passed_days_work   # 휴동일 제외
        rate = round(done_cnt / max_possible * 100, 1) if max_possible else 0
        dept_data.append({
            'dept': dept_nm, 'eq_cnt': eq_cnt,
            'done_cnt': done_cnt, 'max_possible': max_possible, 'rate': rate,
        })
    dept_data.sort(key=lambda x: -x['rate'])

    # 설비별 집계
    eq_data = []
    for r in all_eq_rows:
        done_days = eq_done_days[r['id']]
        rate = round(done_days / passed_days_work * 100, 1) if passed_days_work else 0
        eq_data.append({
            'id': r['id'], 'name': r['name'],
            'dept': r['department'] or '-', 'location': r['location'] or '-',
            'done_days': done_days, 'passed_days': passed_days_work, 'rate': rate,
        })
    eq_data.sort(key=lambda x: (-x['rate'], x['name']))

    # ── 연간 월별 점검율 집계 (주말 제외) ────────────────────────────────────
    yearly_by_month = _dd(set)   # month_num → set of (eq_id, day_str)
    for r in yearly_insp_raw:
        ds   = str(r['day'])
        prts = ds.split('-')
        m_n, d_n = int(prts[1]), int(prts[2])
        try:
            if datetime(year, m_n, d_n).weekday() < 5:  # 평일만
                yearly_by_month[m_n].add((r['equipment_id'], ds))
        except Exception:
            pass

    yearly_labels = [f'{m}월' for m in range(1, 13)]
    yearly_values = []
    for m in range(1, 13):
        days_in_m = calendar.monthrange(year, m)[1]
        if year < now.year or (year == now.year and m < now.month):
            last_day = days_in_m
        elif year == now.year and m == now.month:
            last_day = now.day
        else:
            yearly_values.append(None)   # 미래 월
            continue
        work_days_m = sum(
            1 for d in range(1, last_day + 1)
            if datetime(year, m, d).weekday() < 5
        )
        if work_days_m == 0 or total_eq == 0:
            yearly_values.append(0)
        else:
            done_m = len(yearly_by_month[m])
            yearly_values.append(round(done_m / (work_days_m * total_eq) * 100, 1))

    return render_template('monitoring.html',
        year=year, month=month, days_in_month=days_in_month,
        passed_days=passed_days, passed_days_work=passed_days_work,
        total_eq=total_eq,
        today_done=today_done, today_rate=today_rate,
        today_is_idle=today_is_idle,
        avg_rate=avg_rate,
        idle_days_count=len(idle_days_set),
        idle_days_list=_json.dumps(sorted(idle_days_set)),
        chart_labels=_json.dumps(chart_labels, ensure_ascii=False),
        chart_values=_json.dumps(chart_values),
        yearly_labels=_json.dumps(yearly_labels, ensure_ascii=False),
        yearly_values=_json.dumps(yearly_values),
        dept_data=dept_data, eq_data=eq_data,
        now_year=now.year, now_month=now.month,
        current_dept=current_dept, all_teams=TEAMS,
    )


# ── 전체 설비 리스트 ──────────────────────────────────────────────────────────
@app.route('/equipment-list')
@login_required
def equipment_list():
    conn = get_db()
    dept_sql, dept_params, current_dept = _dept_filter(conn)
    today_cmp = f"({conn.date_col('latest.inspected_at')} = {conn.today})"
    equipments = conn.execute(f'''
        SELECT e.*,
               latest.result        AS last_result,
               latest.status        AS last_status,
               latest.inspected_at  AS last_inspected,
               u.name               AS inspector_name,
               a.name               AS approver_name,
               {today_cmp}          AS inspected_today
        FROM equipment e
        LEFT JOIN inspections latest ON latest.id = (
            SELECT id FROM inspections WHERE equipment_id = e.id ORDER BY inspected_at DESC LIMIT 1
        )
        LEFT JOIN users u ON latest.inspector_id = u.id
        LEFT JOIN users a ON e.approver_id = a.id
        WHERE 1=1 {dept_sql}
        ORDER BY e.name
    ''', dept_params).fetchall()
    conn.close()
    return render_template('equipment_list.html', equipments=equipments,
                           current_dept=current_dept, all_teams=TEAMS)


# ── 월별 점검결과 HTML 페이지 ─────────────────────────────────────────────────
@app.route('/monthly/<int:eq_id>')
@login_required
def monthly_results(eq_id):
    now   = now_kst()
    year  = int(request.args.get('year',  now.year))
    month = int(request.args.get('month', now.month))
    ym    = f"{year}-{month:02d}"

    conn = get_db()
    eq = conn.execute('''
        SELECT e.*, a.name AS approver_name
        FROM equipment e
        LEFT JOIN users a ON e.approver_id = a.id
        WHERE e.id=?
    ''', (eq_id,)).fetchone()
    if not eq:
        conn.close()
        flash('설비를 찾을 수 없습니다.', 'error')
        return redirect(url_for('dashboard'))

    db_items = conn.execute(
        'SELECT * FROM inspection_items WHERE equipment_id=? ORDER BY item_order', (eq_id,)
    ).fetchall()

    tmpl = conn.execute('SELECT * FROM inspection_templates WHERE equipment_id=?', (eq_id,)).fetchone()
    tmpl_rows = json.loads(tmpl['rows']) if tmpl and not db_items else []

    if conn._pg:
        ym_expr = "LEFT(inspected_at, 7)"
    else:
        ym_expr = "strftime('%Y-%m', inspected_at)"

    inspections = conn.execute(f'''
        SELECT i.*, u.name AS inspector_name,
               a.name AS approved_name,
               {conn.date_col("i.inspected_at")} AS insp_date
        FROM inspections i
        JOIN users u ON i.inspector_id = u.id
        LEFT JOIN users a ON i.approved_by = a.id
        WHERE i.equipment_id = ? AND {ym_expr} = ?
        ORDER BY CASE WHEN i.status='승인완료' THEN 0 ELSE 1 END, i.inspected_at DESC
    ''', (eq_id, ym)).fetchall()

    # 날짜별 최선 1건 (승인완료 우선, 이후 최신순) — 먼저 나온 것이 우선
    insp_by_day = {}
    for ins in inspections:
        day = int(str(ins['insp_date']).split('-')[2])
        if day not in insp_by_day:
            insp_by_day[day] = ins

    details_by_insp = {}
    for ins in inspections:
        rows = conn.execute(
            'SELECT * FROM inspection_details WHERE inspection_id=?', (ins['id'],)
        ).fetchall()
        if db_items:
            details_by_insp[ins['id']] = {d['item_id']: d for d in rows if d['item_id']}
        else:
            details_by_insp[ins['id']] = {d['row_index']: d for d in rows}

    # 시리얼 번호 (등록순 rank)
    serial_no = conn.execute(
        'SELECT COUNT(*) AS cnt FROM equipment WHERE id <= ?', (eq_id,)
    ).fetchone()['cnt']

    # 월별 비고
    mn = conn.execute(
        'SELECT notes FROM monthly_notes WHERE equipment_id=? AND year=? AND month=?',
        (eq_id, year, month)
    ).fetchone()
    monthly_note = mn['notes'] if mn else ''

    conn.close()
    days_in_month = calendar.monthrange(year, month)[1]

    # 요일 정보 (0=월 ~ 6=일)
    import datetime as _dt
    weekday_names = ['월', '화', '수', '목', '금', '토', '일']
    day_weekday = {}
    weekend_days = set()
    for d in range(1, days_in_month + 1):
        wd = _dt.date(year, month, d).weekday()  # 0=월, 6=일
        day_weekday[d] = weekday_names[wd]
        if wd >= 5:  # 토(5), 일(6)
            weekend_days.add(d)

    return render_template('monthly_results.html',
        eq=eq, db_items=db_items, tmpl_rows=tmpl_rows,
        insp_by_day=insp_by_day, details_by_insp=details_by_insp,
        year=year, month=month, days_in_month=days_in_month,
        now_year=now.year, now_month=now.month,
        serial_no=f'{serial_no:04d}',
        monthly_note=monthly_note,
        day_weekday=day_weekday, weekend_days=weekend_days)


# ── 월별 비고 저장 ───────────────────────────────────────────────────────────
@app.route('/monthly-note/save', methods=['POST'])
@login_required
def save_monthly_note():
    data = request.get_json()
    eq_id = data.get('equipment_id')
    year  = data.get('year')
    month = data.get('month')
    notes = data.get('notes', '')
    conn  = get_db()
    if conn._pg:
        conn.execute('''
            INSERT INTO monthly_notes (equipment_id, year, month, notes)
            VALUES (%s, %s, %s, %s)
            ON CONFLICT (equipment_id, year, month) DO UPDATE SET notes=EXCLUDED.notes
        ''', (eq_id, year, month, notes))
    else:
        conn.execute('''
            INSERT INTO monthly_notes (equipment_id, year, month, notes) VALUES (?,?,?,?)
            ON CONFLICT(equipment_id, year, month) DO UPDATE SET notes=excluded.notes
        ''', (eq_id, year, month, notes))
    conn.commit()
    conn.close()
    return {'ok': True}


# ── 비고 저장 ─────────────────────────────────────────────────────────────────
@app.route('/inspection/update-notes', methods=['POST'])
@login_required
def update_inspection_notes():
    data  = request.get_json()
    ins_id = data.get('inspection_id')
    notes  = data.get('notes', '')
    conn = get_db()
    conn.execute('UPDATE inspections SET notes=? WHERE id=?', (notes, ins_id))
    conn.commit()
    conn.close()
    return {'ok': True}


# ── 월별 점검결과 엑셀 내보내기 ───────────────────────────────────────────────
@app.route('/export/monthly/<int:eq_id>')
@login_required
def export_monthly(eq_id):
    if not HAS_OPENPYXL:
        flash('openpyxl 패키지가 필요합니다.', 'error')
        return redirect(url_for('inspect', eq_id=eq_id))

    now   = now_kst()
    year  = int(request.args.get('year',  now.year))
    month = int(request.args.get('month', now.month))
    ym    = f"{year}-{month:02d}"

    conn = get_db()
    eq   = conn.execute('SELECT * FROM equipment WHERE id=?', (eq_id,)).fetchone()
    if not eq:
        conn.close()
        flash('설비를 찾을 수 없습니다.', 'error')
        return redirect(url_for('dashboard'))

    tmpl      = conn.execute('SELECT * FROM inspection_templates WHERE equipment_id=?', (eq_id,)).fetchone()
    tmpl_rows = json.loads(tmpl['rows']) if tmpl else []

    # 해당 월 점검 목록
    if conn._pg:
        ym_expr = "LEFT(inspected_at, 7)"
    else:
        ym_expr = "strftime('%Y-%m', inspected_at)"

    inspections = conn.execute(f'''
        SELECT i.*, u.name AS inspector_name,
               {conn.date_col("i.inspected_at")} AS insp_date
        FROM inspections i
        JOIN users u ON i.inspector_id = u.id
        WHERE i.equipment_id = ? AND {ym_expr} = ?
        ORDER BY CASE WHEN i.status='승인완료' THEN 0 ELSE 1 END, i.inspected_at DESC
    ''', (eq_id, ym)).fetchall()

    # 날짜별 최선 1건 (승인완료 우선, 이후 최신순)
    insp_by_day = {}
    for ins in inspections:
        day = int(str(ins['insp_date']).split('-')[2])
        if day not in insp_by_day:
            insp_by_day[day] = ins

    # 항목별 결과 로드
    details_by_insp = {}
    for ins in inspections:
        rows = conn.execute(
            'SELECT * FROM inspection_details WHERE inspection_id=?', (ins['id'],)
        ).fetchall()
        details_by_insp[ins['id']] = {d['row_index']: d for d in rows}

    conn.close()

    # ── Excel 생성 ────────────────────────────────────────────────────────────
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{month:02d}월 점검결과"

    days_in_month = calendar.monthrange(year, month)[1]
    total_cols    = 4 + days_in_month  # 번호+항목+기준+단위 + 일수

    # 스타일 정의
    def fill(hex_):  return PatternFill('solid', fgColor=hex_)
    def font(bold=False, color='111827', size=9):
        return Font(bold=bold, color=color, size=size)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    thin   = Side(style='thin', color='D1D5DB')
    bdr    = Border(left=thin, right=thin, top=thin, bottom=thin)

    ORANGE_FILL  = fill('F97316')
    LIGHT_FILL   = fill('FFF7ED')
    GREEN_FILL   = fill('DCFCE7')
    RED_FILL     = fill('FEE2E2')
    GRAY_FILL    = fill('F3F4F6')
    EMPTY_FILL   = fill('F9FAFB')
    SECTION_FILL = fill('FFEDD5')

    def hdr_cell(row, col, value):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(bold=True, color='FFFFFF', size=9)
        c.fill = ORANGE_FILL
        c.alignment = center
        c.border = bdr
        return c

    # ── 1행: 제목 ────────────────────────────────────────────────────────────
    ws.merge_cells(f'A1:{get_column_letter(total_cols)}1')
    c = ws['A1']
    c.value     = f"{eq['name']}  |  {year}년 {month}월 점검결과표"
    c.font      = Font(bold=True, color='FFFFFF', size=13)
    c.fill      = ORANGE_FILL
    c.alignment = center
    ws.row_dimensions[1].height = 32

    # ── 2행: 설비 정보 ────────────────────────────────────────────────────────
    info = [('설치위치', eq['location'] or '-'), ('부서', eq['department'] or '-')]
    col = 1
    for label, val in info:
        lc = ws.cell(row=2, column=col, value=label)
        lc.font = Font(bold=True, size=9); lc.fill = LIGHT_FILL; lc.alignment = center; lc.border = bdr
        vc = ws.cell(row=2, column=col+1, value=val)
        vc.font = font(size=9); vc.fill = LIGHT_FILL; vc.alignment = left; vc.border = bdr
        col += 2
    # 나머지 셀 채우기
    for c2 in range(col, total_cols+1):
        ws.cell(row=2, column=c2).fill = LIGHT_FILL
        ws.cell(row=2, column=c2).border = bdr
    ws.row_dimensions[2].height = 18

    # ── 4행: 컬럼 헤더 ────────────────────────────────────────────────────────
    HDR_ROW = 4
    for col, h in enumerate(['번호', '점검항목', '판단기준', '단위'], 1):
        hdr_cell(HDR_ROW, col, h)
    for day in range(1, days_in_month+1):
        hdr_cell(HDR_ROW, 4+day, f"{day}")
    ws.row_dimensions[HDR_ROW].height = 20

    # ── 열 너비 ────────────────────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 26
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 7
    for day in range(1, days_in_month+1):
        ws.column_dimensions[get_column_letter(4+day)].width = 5

    # ── 데이터 행 ─────────────────────────────────────────────────────────────
    data_row = HDR_ROW + 1
    item_num = 0

    for row_idx, row in enumerate(tmpl_rows):
        cells = row['cells']

        if not row['is_item']:
            # 섹션 헤더
            ws.merge_cells(f'A{data_row}:{get_column_letter(total_cols)}{data_row}')
            c = ws.cell(row=data_row, column=1, value=' '.join(cells))
            c.font = Font(bold=True, size=9, color='EA580C')
            c.fill = SECTION_FILL; c.alignment = left; c.border = bdr
            for col in range(2, total_cols+1):
                ws.cell(row=data_row, column=col).fill = SECTION_FILL
                ws.cell(row=data_row, column=col).border = bdr
            ws.row_dimensions[data_row].height = 18
            data_row += 1
            continue

        item_num += 1
        # 번호
        c = ws.cell(row=data_row, column=1, value=item_num)
        c.font = font(); c.alignment = center; c.border = bdr

        # 점검항목
        item_name = cells[1] if len(cells) > 1 else (cells[0] if cells else '')
        c = ws.cell(row=data_row, column=2, value=item_name)
        c.font = font(); c.alignment = left; c.border = bdr

        # 판단기준
        c = ws.cell(row=data_row, column=3, value=cells[2] if len(cells) > 2 else '')
        c.font = font(); c.alignment = left; c.border = bdr

        # 단위
        c = ws.cell(row=data_row, column=4, value=cells[3] if len(cells) > 3 else '')
        c.font = font(); c.alignment = center; c.border = bdr

        # 일별 결과
        for day in range(1, days_in_month+1):
            c = ws.cell(row=data_row, column=4+day)
            c.alignment = center; c.border = bdr

            if day in insp_by_day:
                ins = insp_by_day[day]
                detail = details_by_insp.get(ins['id'], {}).get(row_idx)
                if detail:
                    res = detail['result']
                    if res == '정상':
                        c.value = 'O'; c.fill = GREEN_FILL
                        c.font = Font(bold=True, size=9, color='15803D')
                    elif res == '이상':
                        c.value = 'X'; c.fill = RED_FILL
                        c.font = Font(bold=True, size=9, color='DC2626')
                    else:
                        c.value = '-'; c.fill = GRAY_FILL
                        c.font = Font(size=9, color='6B7280')
                else:
                    c.fill = EMPTY_FILL
            else:
                c.fill = EMPTY_FILL

        ws.row_dimensions[data_row].height = 18
        data_row += 1

    # ── 점검자 행 ─────────────────────────────────────────────────────────────
    ws.merge_cells(f'A{data_row}:D{data_row}')
    c = ws.cell(row=data_row, column=1, value='점검자')
    c.font = Font(bold=True, size=9); c.fill = LIGHT_FILL
    c.alignment = center; c.border = bdr

    for day in range(1, days_in_month+1):
        c = ws.cell(row=data_row, column=4+day)
        c.border = bdr; c.alignment = center
        c.font = Font(size=7)
        if day in insp_by_day:
            c.value = insp_by_day[day]['inspector_name']
            c.fill = LIGHT_FILL
        else:
            c.fill = EMPTY_FILL
    ws.row_dimensions[data_row].height = 16
    data_row += 1

    # ── 범례 ──────────────────────────────────────────────────────────────────
    ws.merge_cells(f'A{data_row}:{get_column_letter(total_cols)}{data_row}')
    c = ws.cell(row=data_row, column=1,
                value='※ O: 정상   X: 이상   -: 해당없음   빈칸: 미점검')
    c.font = Font(size=8, italic=True, color='6B7280')
    c.alignment = left

    # ── 출력 ──────────────────────────────────────────────────────────────────
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"{eq['name']}_{year}년{month:02d}월_점검결과.xlsx"
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# ── 관리자: DB 백업 다운로드 ─────────────────────────────────────────────────
@app.route('/admin/backup')
@admin_required
def admin_backup():
    """전체 데이터베이스를 CSV ZIP 파일로 다운로드"""
    import csv, io, zipfile

    TABLES = [
        ('users',
         ['id','name','employee_id','email','phone','team',
          'role','is_admin','is_approved','created_at']),
        ('equipment',
         ['id','name','location','description','qr_code',
          'approver_id','created_at']),
        ('inspections',
         ['id','equipment_id','user_id','status',
          'inspected_at','approved_at','approver_id','note']),
        ('inspection_items',
         ['id','equipment_id','item_name','item_type','options','order_num']),
        ('inspection_templates',
         ['id','equipment_id','filename','uploaded_at']),
        ('inspection_details',
         ['id','inspection_id','item_id','value','note']),
        ('monthly_notes',
         ['id','equipment_id','year','month','note','updated_at']),
        ('password_reset_requests',
         ['id','user_id','requested_at','status','code','expires_at']),
    ]

    conn = get_db()
    buf  = BytesIO()
    now_str = datetime.now().strftime('%Y%m%d_%H%M%S')

    try:
        with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
            # ── 각 테이블 CSV ─────────────────────────────────────────────
            for tbl, cols in TABLES:
                try:
                    rows = conn.execute(
                        f'SELECT {", ".join(cols)} FROM {tbl}'
                    ).fetchall()
                    csv_buf = io.StringIO()
                    w = csv.writer(csv_buf)
                    w.writerow(cols)
                    for row in rows:
                        w.writerow([row[c] for c in cols])
                    zf.writestr(f'{tbl}.csv',
                                csv_buf.getvalue().encode('utf-8-sig').decode('utf-8-sig'))
                except Exception as e:
                    zf.writestr(f'{tbl}_ERROR.txt', str(e))

            # ── 백업 메타 정보 ────────────────────────────────────────────
            meta = (
                f'INTOPS 설비점검 시스템 DB 백업\n'
                f'생성일시: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}\n'
                f'DB종류: {"PostgreSQL" if USE_PG else "SQLite"}\n'
                f'테이블 수: {len(TABLES)}\n'
            )
            zf.writestr('_backup_info.txt', meta)

    finally:
        conn.close()

    buf.seek(0)
    fname = f'intops_backup_{now_str}.zip'
    return send_file(
        buf,
        as_attachment=True,
        download_name=fname,
        mimetype='application/zip'
    )


# ── 내 정보 관리 ─────────────────────────────────────────────────────────────
@app.route('/my-profile', methods=['GET', 'POST'])
@login_required
def my_profile():
    conn = get_db()
    user = conn.execute('SELECT * FROM users WHERE id=?', (session['user_id'],)).fetchone()

    if request.method == 'POST':
        action = request.form.get('action', 'info')

        if action == 'info':
            email = request.form.get('email', '').strip()
            phone = request.form.get('phone', '').strip()
            team  = request.form.get('team', '').strip()
            conn.execute(
                'UPDATE users SET email=?, phone=?, team=? WHERE id=?',
                (email, phone, team, session['user_id'])
            )
            conn.commit()
            conn.close()
            flash('개인정보가 저장되었습니다.', 'success')
            return redirect(url_for('my_profile'))

        elif action == 'password':
            cur_pw   = request.form.get('current_password', '')
            new_pw   = request.form.get('new_password', '')
            new_pw2  = request.form.get('new_password2', '')

            if not check_pw(user['password'], cur_pw):
                conn.close()
                flash('현재 비밀번호가 올바르지 않습니다.', 'error')
                return redirect(url_for('my_profile'))
            if len(new_pw) < 4:
                conn.close()
                flash('새 비밀번호는 4자 이상이어야 합니다.', 'error')
                return redirect(url_for('my_profile'))
            if new_pw != new_pw2:
                conn.close()
                flash('새 비밀번호가 일치하지 않습니다.', 'error')
                return redirect(url_for('my_profile'))

            conn.execute(
                'UPDATE users SET password=? WHERE id=?',
                (hash_pw(new_pw), session['user_id'])
            )
            conn.commit()
            conn.close()
            flash('비밀번호가 변경되었습니다.', 'success')
            return redirect(url_for('my_profile'))

        conn.close()
        return redirect(url_for('my_profile'))

    conn.close()
    return render_template('my_profile.html', user=user, teams=TEAMS)


# ── 이상발생관리 ─────────────────────────────────────────────────────────────

def _save_anomaly_photos(conn, anomaly_id, form):
    """form 에서 photo_data_1~5 를 읽어 anomaly_photos 에 저장. 저장 건수 반환."""
    ph = '%s' if conn._pg else '?'
    saved = 0
    for i in range(1, 6):
        data = form.get(f'photo_data_{i}', '').strip()
        fname = form.get(f'photo_name_{i}', '').strip()
        if data and data.startswith('data:image'):
            conn.execute(
                f'INSERT INTO anomaly_photos (anomaly_id, photo_data, filename) VALUES ({ph},{ph},{ph})',
                (anomaly_id, data, fname)
            )
            saved += 1
    return saved


@app.route('/anomaly-management')
def anomaly_management():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    ph   = '%s' if conn._pg else '?'

    # 필터 파라미터
    f_dept     = request.args.get('dept', '').strip()
    f_resolved = request.args.get('resolved', '').strip()
    f_eq       = request.args.get('eq_id', '').strip()
    f_priority = request.args.get('priority', '').strip()

    equipments = conn.execute('SELECT id, name, department FROM equipment ORDER BY department, name').fetchall()

    where_parts = []
    params      = []

    if f_dept:
        where_parts.append(f'e.department = {ph}')
        params.append(f_dept)
    if f_resolved in ('0', '1'):
        where_parts.append(f'a.is_resolved = {ph}')
        params.append(int(f_resolved))
    if f_eq:
        where_parts.append(f'a.equipment_id = {ph}')
        params.append(int(f_eq))
    if f_priority:
        where_parts.append(f'a.priority = {ph}')
        params.append(f_priority)

    if not session.get('is_admin'):
        my_team = session.get('team', '')
        if my_team and my_team != '관리자':
            where_parts.append(f'e.department = {ph}')
            params.append(my_team)

    where_sql = ('WHERE ' + ' AND '.join(where_parts)) if where_parts else ''

    anomalies = conn.execute(f'''
        SELECT a.*,
               e.name AS eq_name, e.department,
               u.name AS reporter_name,
               r.name AS resolver_name
        FROM equipment_anomalies a
        JOIN equipment e ON e.id = a.equipment_id
        JOIN users u ON u.id = a.reporter_id
        LEFT JOIN users r ON r.id = a.resolved_by
        {where_sql}
        ORDER BY a.occurred_at DESC
    ''', params).fetchall()

    # 각 이상 건 사진 수 조회
    anomaly_ids = [a['id'] for a in anomalies]
    photo_counts = {}
    if anomaly_ids:
        placeholders = ','.join([ph] * len(anomaly_ids))
        rows = conn.execute(
            f'SELECT anomaly_id, COUNT(*) as cnt FROM anomaly_photos WHERE anomaly_id IN ({placeholders}) GROUP BY anomaly_id',
            anomaly_ids
        ).fetchall()
        photo_counts = {r['anomaly_id']: r['cnt'] for r in rows}

    conn.close()
    return render_template('anomaly_management.html',
        anomalies=anomalies, equipments=equipments,
        photo_counts=photo_counts,
        teams=TEAMS,
        f_dept=f_dept, f_resolved=f_resolved,
        f_eq=f_eq, f_priority=f_priority)


@app.route('/anomaly/report', methods=['POST'])
def anomaly_report():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    ph   = '%s' if conn._pg else '?'

    eq_id         = request.form.get('equipment_id', '').strip()
    inspection_id = request.form.get('inspection_id', '').strip() or None
    description   = request.form.get('description', '').strip()
    action_taken  = request.form.get('action_taken', '').strip()
    action_person = request.form.get('action_person', '').strip()
    priority      = request.form.get('priority', '보통').strip()
    is_resolved   = 1 if request.form.get('is_resolved') == '1' else 0
    occurred_at   = request.form.get('occurred_at', '').strip()

    if not eq_id or not description:
        flash('설비와 이상 내용은 필수입니다.', 'error')
        conn.close()
        ref = request.referrer or url_for('anomaly_management')
        return redirect(ref)

    now_str     = now_kst().strftime('%Y-%m-%d %H:%M:%S')
    occ         = occurred_at if occurred_at else now_str
    resolved_at = now_str if is_resolved else None

    anomaly_id = conn.insert(f'''
        INSERT INTO equipment_anomalies
            (equipment_id, inspection_id, reporter_id, occurred_at,
             description, action_taken, action_person, priority,
             is_resolved, resolved_at, resolved_by)
        VALUES ({ph},{ph},{ph},{ph},{ph},{ph},{ph},{ph},{ph},{ph},{ph})
    ''', (eq_id, inspection_id, session['user_id'], occ,
          description, action_taken, action_person, priority,
          is_resolved, resolved_at,
          session['user_id'] if is_resolved else None))

    _save_anomaly_photos(conn, anomaly_id, request.form)
    conn.commit()
    conn.close()
    flash('이상 내용이 등록되었습니다.', 'success')
    ref = request.form.get('next') or request.referrer or url_for('anomaly_management')
    return redirect(ref)


@app.route('/anomaly/<int:anomaly_id>/update', methods=['POST'])
def anomaly_update(anomaly_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    ph   = '%s' if conn._pg else '?'

    action_taken  = request.form.get('action_taken', '').strip()
    action_person = request.form.get('action_person', '').strip()
    priority      = request.form.get('priority', '보통').strip()
    is_resolved   = 1 if request.form.get('is_resolved') == '1' else 0
    now_str       = now_kst().strftime('%Y-%m-%d %H:%M:%S')
    resolved_at   = now_str if is_resolved else None
    resolved_by   = session['user_id'] if is_resolved else None

    conn.execute(f'''
        UPDATE equipment_anomalies
        SET action_taken={ph}, action_person={ph}, priority={ph},
            is_resolved={ph}, resolved_at={ph}, resolved_by={ph}
        WHERE id={ph}
    ''', (action_taken, action_person, priority,
          is_resolved, resolved_at, resolved_by, anomaly_id))

    _save_anomaly_photos(conn, anomaly_id, request.form)
    conn.commit()
    conn.close()
    flash('이상 내용이 업데이트되었습니다.', 'success')
    ref = request.referrer or url_for('anomaly_management')
    return redirect(ref)


@app.route('/anomaly-photo/<int:photo_id>')
def anomaly_photo(photo_id):
    """저장된 base64 사진을 이미지로 반환"""
    if 'user_id' not in session:
        return '', 403
    conn = get_db()
    ph   = '%s' if conn._pg else '?'
    row  = conn.execute(
        f'SELECT photo_data, filename FROM anomaly_photos WHERE id={ph}', (photo_id,)
    ).fetchone()
    conn.close()
    if not row:
        return '', 404
    data = row['photo_data']
    # data:image/jpeg;base64,.... 형식에서 파싱
    if ',' in data:
        header, b64 = data.split(',', 1)
        mime = header.split(':')[1].split(';')[0] if ':' in header else 'image/jpeg'
    else:
        b64  = data
        mime = 'image/jpeg'
    import base64 as _b64
    raw = _b64.b64decode(b64)
    from flask import Response
    return Response(raw, mimetype=mime)


@app.route('/anomaly-photos/<int:anomaly_id>')
def anomaly_photos_list(anomaly_id):
    """특정 이상 건의 사진 목록(id 리스트) JSON 반환"""
    if 'user_id' not in session:
        return {'photos': []}
    conn = get_db()
    ph   = '%s' if conn._pg else '?'
    rows = conn.execute(
        f'SELECT id, filename, created_at FROM anomaly_photos WHERE anomaly_id={ph} ORDER BY id',
        (anomaly_id,)
    ).fetchall()
    conn.close()
    return {'photos': [{'id': r['id'], 'filename': r['filename'], 'created_at': r['created_at']} for r in rows]}


@app.route('/anomaly-photo/<int:photo_id>/delete', methods=['POST'])
def anomaly_photo_delete(photo_id):
    """사진 삭제 (신고자 본인 또는 관리자)"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    conn = get_db()
    ph   = '%s' if conn._pg else '?'
    # anomaly의 reporter 확인
    row = conn.execute(
        f'''SELECT ap.id, ea.reporter_id FROM anomaly_photos ap
            JOIN equipment_anomalies ea ON ea.id=ap.anomaly_id
            WHERE ap.id={ph}''', (photo_id,)
    ).fetchone()
    if row and (session.get('is_admin') or row['reporter_id'] == session['user_id']):
        conn.execute(f'DELETE FROM anomaly_photos WHERE id={ph}', (photo_id,))
        conn.commit()
        flash('사진이 삭제되었습니다.', 'success')
    else:
        flash('삭제 권한이 없습니다.', 'error')
    conn.close()
    ref = request.referrer or url_for('anomaly_management')
    return redirect(ref)


@app.route('/anomaly/<int:anomaly_id>/delete', methods=['POST'])
def anomaly_delete(anomaly_id):
    if 'user_id' not in session or not session.get('is_admin'):
        flash('관리자만 삭제할 수 있습니다.', 'error')
        return redirect(url_for('anomaly_management'))
    conn = get_db()
    ph   = '%s' if conn._pg else '?'
    conn.execute(f'DELETE FROM anomaly_photos WHERE anomaly_id={ph}', (anomaly_id,))
    conn.execute(f'DELETE FROM equipment_anomalies WHERE id={ph}', (anomaly_id,))
    conn.commit()
    conn.close()
    flash('이상 기록이 삭제되었습니다.', 'success')
    return redirect(url_for('anomaly_management'))


# ── 로그아웃 ──────────────────────────────────────────────────────────────────
@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


if __name__ == '__main__':
    init_db()
    port = int(os.environ.get('PORT', 5000))
    debug = not USE_PG
    print("=" * 50)
    print("  INTOPS 설비점검 시스템 시작")
    print(f"  http://localhost:{port} 으로 접속하세요")
    print("  기본 관리자 계정: admin / admin123")
    print("=" * 50)
    app.run(debug=debug, host='0.0.0.0', port=port)
